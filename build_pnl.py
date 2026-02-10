import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

df = pd.read_csv('/sessions/fervent-dazzling-euler/mnt/uploads/AAPL_2023-01-01_to_2023-12-31_data.csv')
df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)
prices = df['Close'].values
n = len(prices)

# ── Recompute signals ──
# MA Crossover
df['SMA20'] = df['Close'].rolling(20).mean()
df['SMA50'] = df['Close'].rolling(50).mean()
df['MA_Pos'] = (df['SMA20'] > df['SMA50']).astype(int)
df['MA_Sig'] = df['MA_Pos'].diff()

# RSI
delta = df['Close'].diff()
gain = delta.clip(lower=0)
loss = -delta.clip(upper=0)
avg_gain = gain.rolling(14, min_periods=14).mean()
avg_loss = loss.rolling(14, min_periods=14).mean()
for i in range(14, n):
    avg_gain.iloc[i] = (avg_gain.iloc[i-1]*13 + gain.iloc[i])/14
    avg_loss.iloc[i] = (avg_loss.iloc[i-1]*13 + loss.iloc[i])/14
rs = avg_gain / avg_loss
df['RSI'] = 100 - (100/(1+rs))

rsi_pos = 0
rsi_positions = [0]*n
for i in range(1, n):
    if pd.isna(df['RSI'].iloc[i]):
        rsi_positions[i] = 0; continue
    if rsi_pos == 0 and df['RSI'].iloc[i] < 30:
        rsi_pos = 1
    elif rsi_pos == 1 and df['RSI'].iloc[i] > 70:
        rsi_pos = 0
    rsi_positions[i] = rsi_pos
df['RSI_Pos'] = rsi_positions

# MACD
df['EMA12'] = df['Close'].ewm(span=12, adjust=False).mean()
df['EMA26'] = df['Close'].ewm(span=26, adjust=False).mean()
df['MACD'] = df['EMA12'] - df['EMA26']
df['MACD_SL'] = df['MACD'].ewm(span=9, adjust=False).mean()
df['MACD_Pos'] = (df['MACD'] > df['MACD_SL']).astype(int)
df['MACD_Sig'] = df['MACD_Pos'].diff()

# ── Extract trades for each strategy ──
STARTING_CAPITAL = 100000.0

def extract_trades(df, pos_col, sig_col, name):
    trades = []
    in_trade = False
    buy_date = buy_price = buy_idx = shares = None
    for i in range(n):
        if not in_trade and df[pos_col].iloc[i] == 1 and (i == 0 or df[pos_col].iloc[i-1] == 0):
            in_trade = True
            buy_idx = i
            buy_date = df['Date'].iloc[i]
            buy_price = df['Close'].iloc[i]
        elif in_trade and (df[pos_col].iloc[i] == 0 and df[pos_col].iloc[i-1] == 1):
            sell_date = df['Date'].iloc[i]
            sell_price = df['Close'].iloc[i]
            trades.append({
                'buy_idx': buy_idx, 'sell_idx': i,
                'buy_date': buy_date, 'sell_date': sell_date,
                'buy_price': buy_price, 'sell_price': sell_price,
            })
            in_trade = False
    if in_trade:
        trades.append({
            'buy_idx': buy_idx, 'sell_idx': n-1,
            'buy_date': buy_date, 'sell_date': df['Date'].iloc[-1],
            'buy_price': buy_price, 'sell_price': df['Close'].iloc[-1],
        })
    return trades

ma_trades = extract_trades(df, 'MA_Pos', 'MA_Sig', 'MA')
rsi_trades = extract_trades(df, 'RSI_Pos', None, 'RSI')
macd_trades = extract_trades(df, 'MACD_Pos', 'MACD_Sig', 'MACD')

# Buy & Hold: single trade
bh_trades = [{
    'buy_idx': 0, 'sell_idx': n-1,
    'buy_date': df['Date'].iloc[0], 'sell_date': df['Date'].iloc[-1],
    'buy_price': df['Close'].iloc[0], 'sell_price': df['Close'].iloc[-1],
}]

# ── Build daily P&L for each strategy ──
def build_daily_pnl(df, pos_col, trades, capital=STARTING_CAPITAL):
    rows = []
    cash = capital
    shares = 0
    position = 'CASH'
    trade_num = 0
    cum_realized = 0.0
    cost_basis = 0.0

    for i in range(n):
        prev_pos = df[pos_col].iloc[i-1] if i > 0 else 0
        curr_pos = df[pos_col].iloc[i]
        action = ''
        trade_pnl = 0.0

        if curr_pos == 1 and prev_pos == 0:
            # BUY
            action = 'BUY'
            trade_num += 1
            shares = int(cash / df['Close'].iloc[i])
            cost_basis = shares * df['Close'].iloc[i]
            cash -= cost_basis
            position = 'LONG'
        elif curr_pos == 0 and prev_pos == 1:
            # SELL
            action = 'SELL'
            proceeds = shares * df['Close'].iloc[i]
            trade_pnl = proceeds - cost_basis
            cum_realized += trade_pnl
            cash += proceeds
            shares = 0
            position = 'CASH'

        mkt_value = shares * df['Close'].iloc[i]
        unrealized = mkt_value - cost_basis if shares > 0 else 0
        total_equity = cash + mkt_value

        rows.append({
            'Date': df['Date'].iloc[i],
            'Close': df['Close'].iloc[i],
            'Action': action,
            'Trade_#': trade_num if action else '',
            'Position': position,
            'Shares': shares,
            'Cost_Basis': cost_basis if shares > 0 else 0,
            'Mkt_Value': mkt_value,
            'Unrealized_PnL': unrealized,
            'Trade_PnL': trade_pnl if action == 'SELL' else '',
            'Cum_Realized_PnL': cum_realized,
            'Cash': cash,
            'Total_Equity': total_equity,
            'Daily_Return_Pct': ((total_equity / rows[-1]['Total_Equity']) - 1)*100 if len(rows) > 0 else 0,
        })
    return pd.DataFrame(rows)

# Build for each strategy
# For Buy & Hold we manually build position column
df['BH_Pos'] = 1

ma_pnl = build_daily_pnl(df, 'MA_Pos', ma_trades)
rsi_pnl = build_daily_pnl(df, 'RSI_Pos', rsi_trades)
macd_pnl = build_daily_pnl(df, 'MACD_Pos', macd_trades)
bh_pnl = build_daily_pnl(df, 'BH_Pos', bh_trades)

# ── Build Excel workbook ──
wb = Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill('solid', fgColor='1a237e')
title_font = Font(name='Arial', bold=True, size=14, color='1a237e')
section_font = Font(name='Arial', bold=True, size=11, color='1a237e')
num_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')
green_font = Font(name='Arial', size=10, color='008000', bold=True)
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
buy_fill = PatternFill('solid', fgColor='E8F5E9')
sell_fill = PatternFill('solid', fgColor='FFEBEE')
total_fill = PatternFill('solid', fgColor='E3F2FD')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
bottom_border = Border(bottom=Side(style='medium', color='1a237e'))

money_fmt = '$#,##0.00'
pct_fmt = '0.00%'
int_fmt = '#,##0'
pnl_fmt = '$#,##0.00;[Red]($#,##0.00);"-"'

def write_summary_sheet(wb):
    ws = wb.active
    ws.title = 'P&L Summary'
    ws.sheet_properties.tabColor = '1a237e'

    ws['A1'] = 'AAPL 2023 — Strategy P&L Summary'
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='1a237e')
    ws['A2'] = f'Starting Capital: $100,000 | Period: Jan 3 – Dec 29, 2023'
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    # Summary table
    headers = ['Metric', 'Buy & Hold', 'MA Crossover (20/50)', 'RSI (14)', 'MACD (12/26/9)']
    r = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    strategies = [
        ('Buy & Hold', bh_pnl, bh_trades),
        ('MA Crossover', ma_pnl, ma_trades),
        ('RSI', rsi_pnl, rsi_trades),
        ('MACD', macd_pnl, macd_trades),
    ]

    metrics = []
    for name, pnl_df, trades in strategies:
        final_eq = pnl_df['Total_Equity'].iloc[-1]
        total_ret = (final_eq / STARTING_CAPITAL - 1)
        cum_real = pnl_df['Cum_Realized_PnL'].iloc[-1]
        unreal = pnl_df['Unrealized_PnL'].iloc[-1]
        total_pnl = cum_real + unreal
        num_trades = len(trades)
        winners = sum(1 for t in trades if t['sell_price'] > t['buy_price'])
        losers = num_trades - winners
        win_rate = winners/num_trades if num_trades > 0 else 0
        daily_rets = pnl_df['Daily_Return_Pct']/100
        vol = daily_rets.std() * np.sqrt(252) if daily_rets.std() > 0 else 0
        sharpe = (daily_rets.mean() / daily_rets.std()) * np.sqrt(252) if daily_rets.std() > 0 else 0
        peak = pnl_df['Total_Equity'].cummax()
        dd = ((pnl_df['Total_Equity'] - peak) / peak).min()
        best_trade = max([t['sell_price']/t['buy_price']-1 for t in trades]) if trades else 0
        worst_trade = min([t['sell_price']/t['buy_price']-1 for t in trades]) if trades else 0
        avg_trade = np.mean([t['sell_price']/t['buy_price']-1 for t in trades]) if trades else 0
        days_in = pnl_df[pnl_df['Position']=='LONG'].shape[0]
        exposure = days_in / n

        metrics.append({
            'Starting Capital': STARTING_CAPITAL,
            'Ending Equity': final_eq,
            'Total P&L ($)': total_pnl,
            'Total Return (%)': total_ret,
            'Realized P&L ($)': cum_real,
            'Unrealized P&L ($)': unreal,
            'Number of Trades': num_trades,
            'Winning Trades': winners,
            'Losing Trades': losers,
            'Win Rate (%)': win_rate,
            'Best Trade (%)': best_trade,
            'Worst Trade (%)': worst_trade,
            'Avg Trade Return (%)': avg_trade,
            'Annualized Volatility (%)': vol,
            'Sharpe Ratio': sharpe,
            'Max Drawdown (%)': dd,
            'Market Exposure (%)': exposure,
            'Days in Market': days_in,
            'Days in Cash': n - days_in,
        })

    row_formats = {
        'Starting Capital': money_fmt,
        'Ending Equity': money_fmt,
        'Total P&L ($)': pnl_fmt,
        'Total Return (%)': '0.00%',
        'Realized P&L ($)': pnl_fmt,
        'Unrealized P&L ($)': pnl_fmt,
        'Number of Trades': '0',
        'Winning Trades': '0',
        'Losing Trades': '0',
        'Win Rate (%)': '0.0%',
        'Best Trade (%)': '0.00%',
        'Worst Trade (%)': '0.00%',
        'Avg Trade Return (%)': '0.00%',
        'Annualized Volatility (%)': '0.00%',
        'Sharpe Ratio': '0.00',
        'Max Drawdown (%)': '0.00%',
        'Market Exposure (%)': '0.0%',
        'Days in Market': '0',
        'Days in Cash': '0',
    }

    for i, (metric_name, fmt) in enumerate(row_formats.items()):
        row = r + 1 + i
        cell = ws.cell(row=row, column=1, value=metric_name)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = thin_border

        for j, m in enumerate(metrics):
            cell = ws.cell(row=row, column=j+2, value=m[metric_name])
            cell.number_format = fmt
            cell.font = num_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            if metric_name == 'Total P&L ($)':
                cell.font = green_font if m[metric_name] > 0 else red_font
            if metric_name == 'Total Return (%)':
                cell.font = green_font if m[metric_name] > 0 else red_font
            if metric_name == 'Ending Equity':
                cell.fill = total_fill

    # Highlight best in each row
    highlight_fill = PatternFill('solid', fgColor='C8E6C9')
    for i, metric_name in enumerate(row_formats.keys()):
        row = r + 1 + i
        if metric_name in ['Starting Capital', 'Days in Cash']:
            continue
        vals = [metrics[j][metric_name] for j in range(4)]
        if metric_name in ['Losing Trades', 'Annualized Volatility (%)', 'Max Drawdown (%)', 'Worst Trade (%)']:
            best_idx = vals.index(max(vals))  # least negative
        elif metric_name in ['Days in Market']:
            continue
        else:
            best_idx = vals.index(max(vals))
        ws.cell(row=row, column=best_idx+2).fill = highlight_fill

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Trade detail section
    trade_row = r + len(row_formats) + 3
    ws.cell(row=trade_row, column=1, value='TRADE-BY-TRADE P&L DETAIL').font = Font(name='Arial', bold=True, size=13, color='1a237e')
    trade_row += 1

    for strat_name, trades_list, pnl_df in [
        ('Buy & Hold', bh_trades, bh_pnl),
        ('MA Crossover (20/50)', ma_trades, ma_pnl),
        ('MACD (12/26/9)', macd_trades, macd_pnl),
    ]:
        if not trades_list:
            continue
        trade_row += 1
        ws.cell(row=trade_row, column=1, value=strat_name).font = section_font
        trade_row += 1

        trade_headers = ['Trade #', 'Buy Date', 'Buy Price', 'Sell Date', 'Sell Price',
                         'Shares', 'Gross P&L ($)', 'Return (%)', 'Holding Days', 'Result']
        for c, h in enumerate(trade_headers, 1):
            cell = ws.cell(row=trade_row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for t_idx, t in enumerate(trades_list):
            trade_row += 1
            shares = int(STARTING_CAPITAL / t['buy_price'])  # simplified
            gross_pnl = shares * (t['sell_price'] - t['buy_price'])
            ret = (t['sell_price'] / t['buy_price']) - 1
            days = (t['sell_date'] - t['buy_date']).days

            row_data = [
                t_idx + 1,
                t['buy_date'].strftime('%Y-%m-%d'),
                t['buy_price'],
                t['sell_date'].strftime('%Y-%m-%d'),
                t['sell_price'],
                shares,
                gross_pnl,
                ret,
                days,
                'WIN' if ret > 0 else 'LOSS'
            ]
            fmts = ['0', None, money_fmt, None, money_fmt, int_fmt, pnl_fmt, '0.00%', '0', None]

            for c, (val, fmt) in enumerate(zip(row_data, fmts), 1):
                cell = ws.cell(row=trade_row, column=c, value=val)
                if fmt: cell.number_format = fmt
                cell.font = num_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

                if c == 10:
                    cell.font = green_font if val == 'WIN' else red_font

            fill = buy_fill if ret > 0 else sell_fill
            for c in range(1, 11):
                ws.cell(row=trade_row, column=c).fill = fill

        trade_row += 1

    ws.column_dimensions[get_column_letter(7)].width = 16
    ws.column_dimensions[get_column_letter(8)].width = 14
    ws.column_dimensions[get_column_letter(10)].width = 10

    ws.freeze_panes = 'A5'
    return ws

def write_daily_sheet(wb, sheet_name, pnl_df, tab_color):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    ws['A1'] = f'{sheet_name} — Daily P&L'
    ws['A1'].font = title_font
    ws['A2'] = f'Starting Capital: $100,000'
    ws['A2'].font = Font(name='Arial', size=9, color='666666')

    headers = ['Date', 'Close', 'Action', 'Trade #', 'Position', 'Shares',
               'Cost Basis', 'Mkt Value', 'Unrealized P&L', 'Trade P&L',
               'Cum Realized P&L', 'Cash', 'Total Equity', 'Daily Return %']

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    col_fmts = [None, money_fmt, None, None, None, int_fmt,
                money_fmt, money_fmt, pnl_fmt, pnl_fmt,
                pnl_fmt, money_fmt, money_fmt, '0.00%']

    for i, row in pnl_df.iterrows():
        r = i + 4
        values = [
            row['Date'].strftime('%Y-%m-%d'), row['Close'], row['Action'],
            row['Trade_#'], row['Position'], row['Shares'],
            row['Cost_Basis'], row['Mkt_Value'], row['Unrealized_PnL'],
            row['Trade_PnL'] if row['Trade_PnL'] != '' else '',
            row['Cum_Realized_PnL'], row['Cash'], row['Total_Equity'],
            row['Daily_Return_Pct'] / 100
        ]
        for c, (val, fmt) in enumerate(zip(values, col_fmts), 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fmt and val != '': cell.number_format = fmt
            cell.font = num_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        if row['Action'] == 'BUY':
            for c in range(1, 15):
                ws.cell(row=r, column=c).fill = buy_fill
                if c == 3: ws.cell(row=r, column=c).font = green_font
        elif row['Action'] == 'SELL':
            for c in range(1, 15):
                ws.cell(row=r, column=c).fill = sell_fill
                if c == 3: ws.cell(row=r, column=c).font = red_font

    # Summary row at bottom
    last_row = n + 4
    ws.cell(row=last_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=last_row, column=13, value=f'=N{last_row-1}')
    ws.cell(row=last_row, column=13).font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=last_row, column=13).number_format = money_fmt

    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(headers[c-1]) + 4)
    ws.column_dimensions['A'].width = 12

    ws.freeze_panes = 'A4'
    return ws

# Build workbook
write_summary_sheet(wb)
write_daily_sheet(wb, 'Buy & Hold Daily', bh_pnl, '4CAF50')
write_daily_sheet(wb, 'MA Crossover Daily', ma_pnl, '2196F3')
write_daily_sheet(wb, 'RSI Daily', rsi_pnl, 'AB47BC')
write_daily_sheet(wb, 'MACD Daily', macd_pnl, 'FF7043')

out_path = '/sessions/fervent-dazzling-euler/mnt/outputs/AAPL_2023_PnL_Analysis.xlsx'
wb.save(out_path)
print(f"Workbook saved: {out_path}")
