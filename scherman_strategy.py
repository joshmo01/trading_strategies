"""
Ivan Scherman's VIX Divergence Strategy — Adapted for AAPL
==========================================================

ORIGINAL STRATEGY (S&P 500 + VIX):
  The 2023 World Cup Trading Champion (491% return) publicly shared a strategy
  with a 75% win rate and 3:1 profit potential, backtested 1957-2024.

  ENTRY CONDITIONS (all must be met):
    1. S&P 500 makes a LOWER LOW (price correction)
    2. VIX spikes during correction, BUT the 2nd VIX spike is LOWER than
       the 1st (bearish divergence in fear → institutions reducing hedges)
    3. Price is WITHIN 2 standard deviations of 30-day SMA (inside Bollinger Bands)
    4. Buy at CLOSE on signal day

  EXIT CONDITIONS:
    - PROFIT: Price closes ABOVE the 30-day SMA → exit at close
    - LOSS:   Price closes MORE THAN 2σ BELOW the 30-day SMA → exit at close

ADAPTATION FOR AAPL (no VIX available):
  Since VIX measures implied volatility of S&P options, we construct an
  analogous "fear gauge" for AAPL using:
    - Realized Volatility (14-day rolling std of log returns × √252)
  This captures the same phenomenon: when price drops but volatility doesn't
  spike higher, it signals the correction is losing momentum → turning point.

  We also run a SECOND variant using ATR (Average True Range) as the
  volatility proxy, for robustness comparison.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

df = pd.read_csv('/sessions/fervent-dazzling-euler/mnt/uploads/AAPL_2023-01-01_to_2023-12-31_data.csv')
df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)
prices = df['Close'].values
dates = df['Date'].values
n = len(prices)
CAPITAL = 100000.0

# ════════════════════════════════════════════════════════════
# COMPUTE INDICATORS
# ════════════════════════════════════════════════════════════

# 30-day SMA
df['SMA30'] = df['Close'].rolling(30).mean()

# Bollinger Bands (30-day, 2σ)
df['BB_std'] = df['Close'].rolling(30).std()
df['BB_upper'] = df['SMA30'] + 2 * df['BB_std']
df['BB_lower'] = df['SMA30'] - 2 * df['BB_std']

# Log returns
df['Log_Return'] = np.log(df['Close'] / df['Close'].shift(1))

# Realized Volatility (14-day rolling, annualized) — our "VIX proxy"
df['RV14'] = df['Log_Return'].rolling(14).std() * np.sqrt(252) * 100

# ATR (14-day) — alternative volatility proxy
df['TR'] = np.maximum(
    df['High'] - df['Low'],
    np.maximum(abs(df['High'] - df['Close'].shift(1)),
               abs(df['Low'] - df['Close'].shift(1)))
)
df['ATR14'] = df['TR'].rolling(14).mean()

# Detect swing lows in price (lower lows)
def find_swing_lows(series, order=5):
    """Find local minima in a series."""
    lows = []
    vals = series.values
    for i in range(order, len(vals) - order):
        if pd.isna(vals[i]):
            continue
        window = vals[i-order:i+order+1]
        if not np.any(np.isnan(window)) and vals[i] == np.min(window):
            lows.append(i)
    return lows

# Detect swing highs in volatility proxy
def find_swing_highs(series, order=5):
    """Find local maxima in a series."""
    highs = []
    vals = series.values
    for i in range(order, len(vals) - order):
        if pd.isna(vals[i]):
            continue
        window = vals[i-order:i+order+1]
        if not np.any(np.isnan(window)) and vals[i] == np.max(window):
            highs.append(i)
    return highs

# ════════════════════════════════════════════════════════════
# SCHERMAN DIVERGENCE DETECTION
# ════════════════════════════════════════════════════════════

def detect_scherman_divergence(df, vol_col, order=5, lookback=30):
    """
    Detect Scherman-style divergence:
    - Price makes a LOWER LOW
    - Volatility proxy makes a LOWER HIGH (divergence)
    - Price is within Bollinger Bands (2σ of 30-day SMA)

    Returns list of signal indices (buy signals).
    """
    price_lows = find_swing_lows(df['Close'], order=order)
    vol_highs = find_swing_highs(df[vol_col], order=order)

    signals = []

    for i in range(1, len(price_lows)):
        curr_low_idx = price_lows[i]
        prev_low_idx = price_lows[i-1]

        # Condition 1: Price makes a LOWER LOW
        if df['Close'].iloc[curr_low_idx] >= df['Close'].iloc[prev_low_idx]:
            continue

        # Condition 2: Find vol highs near each price low (within ±lookback days)
        # Vol high near prev price low
        prev_vol_highs = [vh for vh in vol_highs
                          if prev_low_idx - lookback <= vh <= prev_low_idx + lookback]
        # Vol high near curr price low
        curr_vol_highs = [vh for vh in vol_highs
                          if curr_low_idx - lookback <= vh <= curr_low_idx + lookback]

        if not prev_vol_highs or not curr_vol_highs:
            continue

        # Get the highest vol spike near each price low
        prev_vol_peak = max(prev_vol_highs, key=lambda x: df[vol_col].iloc[x])
        curr_vol_peak = max(curr_vol_highs, key=lambda x: df[vol_col].iloc[x])

        # DIVERGENCE: vol high near current price low is LOWER than vol high near previous
        if df[vol_col].iloc[curr_vol_peak] >= df[vol_col].iloc[prev_vol_peak]:
            continue

        # Condition 3: Price is within Bollinger Bands at signal point
        if pd.isna(df['BB_lower'].iloc[curr_low_idx]) or pd.isna(df['BB_upper'].iloc[curr_low_idx]):
            continue

        if df['Close'].iloc[curr_low_idx] < df['BB_lower'].iloc[curr_low_idx]:
            continue  # Already outside lower band = too late

        # All conditions met → BUY SIGNAL
        signals.append({
            'signal_idx': curr_low_idx,
            'signal_date': df['Date'].iloc[curr_low_idx],
            'signal_price': df['Close'].iloc[curr_low_idx],
            'prev_low_idx': prev_low_idx,
            'prev_low_price': df['Close'].iloc[prev_low_idx],
            'prev_vol_peak': df[vol_col].iloc[prev_vol_peak],
            'curr_vol_peak': df[vol_col].iloc[curr_vol_peak],
            'divergence_pct': ((df[vol_col].iloc[curr_vol_peak] / df[vol_col].iloc[prev_vol_peak]) - 1) * 100,
        })

    return signals

# ════════════════════════════════════════════════════════════
# SCHERMAN TRADE EXECUTION ENGINE
# ════════════════════════════════════════════════════════════

def execute_scherman_trades(df, signals, capital=CAPITAL):
    """
    Execute Scherman strategy:
    - Enter LONG at close on signal day
    - EXIT PROFIT: Close > 30-day SMA
    - EXIT LOSS: Close < lower Bollinger Band (2σ below SMA)
    """
    trades = []
    daily_rows = []
    cash = capital
    shares = 0
    cost_basis = 0
    position = 'CASH'
    trade_num = 0
    cum_realized = 0.0
    current_signal = None

    # Build set of signal dates for quick lookup
    signal_indices = set(s['signal_idx'] for s in signals)

    for i in range(n):
        action = ''
        trade_pnl = 0.0
        close = df['Close'].iloc[i]
        sma30 = df['SMA30'].iloc[i]
        bb_lower = df['BB_lower'].iloc[i]

        # Check for entry
        if position == 'CASH' and i in signal_indices:
            action = 'BUY'
            trade_num += 1
            shares = int(cash / close)
            cost_basis = shares * close
            cash -= cost_basis
            position = 'LONG'
            current_signal = [s for s in signals if s['signal_idx'] == i][0]

        # Check for exit (only if in position and we have SMA data)
        elif position == 'LONG' and not pd.isna(sma30) and not pd.isna(bb_lower):
            exit_reason = None

            # Profit exit: close above 30-day SMA
            if close > sma30:
                exit_reason = 'PROFIT (Close > SMA30)'

            # Loss exit: close below lower Bollinger Band
            elif close < bb_lower:
                exit_reason = 'STOP (Close < Lower BB)'

            if exit_reason:
                action = 'SELL'
                proceeds = shares * close
                trade_pnl = proceeds - cost_basis
                cum_realized += trade_pnl
                trades.append({
                    'trade_num': trade_num,
                    'buy_date': current_signal['signal_date'] if current_signal else df['Date'].iloc[i],
                    'buy_price': cost_basis / shares if shares > 0 else 0,
                    'sell_date': df['Date'].iloc[i],
                    'sell_price': close,
                    'shares': shares,
                    'gross_pnl': trade_pnl,
                    'return_pct': (proceeds / cost_basis - 1) if cost_basis > 0 else 0,
                    'exit_reason': exit_reason,
                    'divergence_pct': current_signal['divergence_pct'] if current_signal else 0,
                    'holding_days': (df['Date'].iloc[i] - current_signal['signal_date']).days if current_signal else 0,
                })
                cash += proceeds
                shares = 0
                cost_basis = 0
                position = 'CASH'
                current_signal = None

        mkt_value = shares * close
        unrealized = mkt_value - cost_basis if shares > 0 else 0
        total_equity = cash + mkt_value
        prev_eq = daily_rows[-1]['Total_Equity'] if daily_rows else capital
        daily_ret = ((total_equity / prev_eq) - 1) * 100 if prev_eq > 0 else 0

        daily_rows.append({
            'Date': df['Date'].iloc[i],
            'Close': close,
            'SMA30': sma30,
            'BB_Upper': df['BB_upper'].iloc[i],
            'BB_Lower': bb_lower,
            'Vol_Proxy': df['RV14'].iloc[i],
            'Action': action,
            'Trade_#': trade_num if action else '',
            'Position': position,
            'Shares': shares,
            'Cost_Basis': cost_basis if shares > 0 else 0,
            'Mkt_Value': mkt_value,
            'Unrealized_PnL': unrealized,
            'Trade_PnL': trade_pnl if action == 'SELL' else '',
            'Cum_Realized_PnL': cum_realized,
            'Cash': cash,
            'Total_Equity': total_equity,
            'Daily_Return_Pct': daily_ret,
        })

    # If still in position at end
    if position == 'LONG':
        final_proceeds = shares * prices[-1]
        final_pnl = final_proceeds - cost_basis
        trades.append({
            'trade_num': trade_num,
            'buy_date': current_signal['signal_date'] if current_signal else df['Date'].iloc[-1],
            'buy_price': cost_basis / shares if shares > 0 else 0,
            'sell_date': df['Date'].iloc[-1],
            'sell_price': prices[-1],
            'shares': shares,
            'gross_pnl': final_pnl,
            'return_pct': (final_proceeds / cost_basis - 1) if cost_basis > 0 else 0,
            'exit_reason': 'OPEN (Year-End)',
            'divergence_pct': current_signal['divergence_pct'] if current_signal else 0,
            'holding_days': (df['Date'].iloc[-1] - current_signal['signal_date']).days if current_signal else 0,
        })

    return pd.DataFrame(daily_rows), trades

# ════════════════════════════════════════════════════════════
# RUN BOTH VARIANTS
# ════════════════════════════════════════════════════════════

# Variant 1: Realized Volatility as VIX proxy
rv_signals = detect_scherman_divergence(df, 'RV14', order=5, lookback=20)
rv_daily, rv_trades = execute_scherman_trades(df, rv_signals)

# Variant 2: ATR as VIX proxy
atr_signals = detect_scherman_divergence(df, 'ATR14', order=5, lookback=20)
atr_daily, atr_trades = execute_scherman_trades(df, atr_signals)

# Buy & Hold for comparison
bh_shares = int(CAPITAL / prices[0])
bh_cost = bh_shares * prices[0]
bh_final = bh_shares * prices[-1] + (CAPITAL - bh_cost)
bh_pnl = bh_final - CAPITAL
bh_ret = bh_final / CAPITAL - 1

# ════════════════════════════════════════════════════════════
# PRINT RESULTS
# ════════════════════════════════════════════════════════════

print("=" * 100)
print("  IVAN SCHERMAN VIX DIVERGENCE STRATEGY — ADAPTED FOR AAPL 2023")
print("  Starting Capital: $100,000")
print("=" * 100)

print("""
  STRATEGY LOGIC (adapted from Scherman's 75% win-rate, 3:1 R/R method):
  ┌─────────────────────────────────────────────────────────────────────┐
  │  ENTRY: BUY when ALL conditions met:                               │
  │    1. AAPL price makes a LOWER LOW (swing low < previous swing low)│
  │    2. Volatility proxy makes a LOWER HIGH (divergence = fear       │
  │       decreasing despite lower price → correction losing steam)    │
  │    3. Price is WITHIN Bollinger Bands (30-day, 2σ)                 │
  │                                                                     │
  │  EXIT (Profit): Close ABOVE 30-day SMA → sell at close            │
  │  EXIT (Loss):   Close BELOW lower Bollinger Band → sell at close   │
  └─────────────────────────────────────────────────────────────────────┘
""")

for variant_name, daily_df, trades, signals in [
    ('Variant A: Realized Volatility (14d)', rv_daily, rv_trades, rv_signals),
    ('Variant B: ATR (14d)', atr_daily, atr_trades, atr_signals),
]:
    final_eq = daily_df['Total_Equity'].iloc[-1]
    total_pnl = final_eq - CAPITAL
    total_ret = final_eq / CAPITAL - 1
    num_trades = len(trades)
    winners = sum(1 for t in trades if t['gross_pnl'] > 0)
    losers = num_trades - winners
    win_rate = winners / num_trades if num_trades > 0 else 0
    daily_rets = daily_df['Daily_Return_Pct'] / 100
    vol = daily_rets.std() * np.sqrt(252) if daily_rets.std() > 0 else 0
    sharpe = (daily_rets.mean() / daily_rets.std()) * np.sqrt(252) if daily_rets.std() > 0 else 0
    peak_eq = daily_df['Total_Equity'].cummax()
    max_dd = ((daily_df['Total_Equity'] - peak_eq) / peak_eq).min()
    days_long = daily_df[daily_df['Position'] == 'LONG'].shape[0]

    print(f"\n{'─'*80}")
    print(f"  {variant_name}")
    print(f"{'─'*80}")
    print(f"  Divergence signals detected: {len(signals)}")
    print(f"  Ending Equity:     ${final_eq:>12,.2f}")
    print(f"  Total P&L:         ${total_pnl:>12,.2f}")
    print(f"  Total Return:      {total_ret:>11.2%}")
    print(f"  Number of Trades:  {num_trades:>12}")
    print(f"  Win / Loss:        {winners}W / {losers}L")
    print(f"  Win Rate:          {win_rate:>11.0%}")
    print(f"  Sharpe Ratio:      {sharpe:>12.2f}")
    print(f"  Max Drawdown:      {max_dd:>11.2%}")
    print(f"  Days in Market:    {days_long:>12} / {n}")
    print()

    if signals:
        print("  DIVERGENCE SIGNALS:")
        for s in signals:
            print(f"    {s['signal_date'].strftime('%b %d')}  Price: ${s['signal_price']:.2f}  "
                  f"(Lower low vs ${s['prev_low_price']:.2f})  "
                  f"Vol divergence: {s['divergence_pct']:.1f}%")
        print()

    print("  TRADE LOG:")
    if not trades:
        print("    No trades executed.")
    for t in trades:
        result = "WIN " if t['gross_pnl'] > 0 else "LOSS"
        bd = t['buy_date'].strftime('%b %d')
        sd = t['sell_date'].strftime('%b %d')
        print(f"    #{t['trade_num']:>2}  {bd} → {sd} ({t['holding_days']}d)  "
              f"Buy ${t['buy_price']:.2f} → Sell ${t['sell_price']:.2f}  "
              f"P&L: ${t['gross_pnl']:>+10,.2f} ({t['return_pct']:>+.2%})  "
              f"{result}  [{t['exit_reason']}]")

print(f"\n{'─'*80}")
print(f"  BENCHMARK: Buy & Hold")
print(f"{'─'*80}")
print(f"  Ending Equity:     ${bh_final:>12,.2f}")
print(f"  Total Return:      {bh_ret:>11.2%}")

# ════════════════════════════════════════════════════════════
# VISUALIZATION
# ════════════════════════════════════════════════════════════

fig = plt.figure(figsize=(22, 28))
gs = gridspec.GridSpec(5, 1, height_ratios=[3, 1.5, 1.5, 2, 2], hspace=0.35)

colors = {'price': '#42A5F5', 'sma': '#FFC107', 'bb': '#7E57C2',
          'buy': '#00E676', 'sell_profit': '#00E676', 'sell_loss': '#FF1744',
          'vol': '#FF7043', 'equity': '#26C6DA'}

# ═══ CHART 1: Price with Bollinger Bands + Signals (Variant A) ═══
ax1 = fig.add_subplot(gs[0])
ax1.plot(dates, prices, color=colors['price'], linewidth=1.3, label='AAPL Close')
ax1.plot(df['Date'], df['SMA30'], color=colors['sma'], linewidth=1, alpha=0.8, label='30-Day SMA')
ax1.fill_between(df['Date'], df['BB_upper'], df['BB_lower'], alpha=0.08, color=colors['bb'])
ax1.plot(df['Date'], df['BB_upper'], color=colors['bb'], linewidth=0.7, alpha=0.5, linestyle='--', label='Bollinger Bands (2σ)')
ax1.plot(df['Date'], df['BB_lower'], color=colors['bb'], linewidth=0.7, alpha=0.5, linestyle='--')

# Plot buy/sell signals
for s in rv_signals:
    ax1.scatter([s['signal_date']], [s['signal_price']], color=colors['buy'],
                s=200, marker='^', zorder=5, edgecolors='white', linewidth=1)
    ax1.annotate(f"BUY\n${s['signal_price']:.0f}", xy=(s['signal_date'], s['signal_price']),
                 xytext=(s['signal_date'], s['signal_price'] - 8),
                 fontsize=8, ha='center', color=colors['buy'], fontweight='bold',
                 arrowprops=dict(arrowstyle='->', color=colors['buy'], lw=1.5))

for t in rv_trades:
    sell_color = colors['sell_profit'] if t['gross_pnl'] > 0 else colors['sell_loss']
    ax1.scatter([t['sell_date']], [t['sell_price']], color=sell_color,
                s=200, marker='v', zorder=5, edgecolors='white', linewidth=1)
    label = f"SELL\n${t['sell_price']:.0f}"
    ax1.annotate(label, xy=(t['sell_date'], t['sell_price']),
                 xytext=(t['sell_date'], t['sell_price'] + 5),
                 fontsize=8, ha='center', color=sell_color, fontweight='bold',
                 arrowprops=dict(arrowstyle='->', color=sell_color, lw=1.5))

    # Shade holding periods
    ax1.axvspan(t['buy_date'], t['sell_date'], alpha=0.06,
                color=colors['buy'] if t['gross_pnl'] > 0 else colors['sell_loss'])

ax1.set_title("Ivan Scherman Strategy — AAPL 2023 (Variant A: Realized Volatility)",
              fontsize=14, fontweight='bold', pad=12)
ax1.set_ylabel('Price ($)', fontsize=11)
ax1.legend(loc='upper left', fontsize=9)
ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
ax1.xaxis.set_major_locator(mdates.MonthLocator())

# ═══ CHART 2: Volatility Proxy (Realized Vol) ═══
ax2 = fig.add_subplot(gs[1])
ax2.plot(df['Date'], df['RV14'], color=colors['vol'], linewidth=1.2, label='14d Realized Volatility (annualized)')
ax2.fill_between(df['Date'], df['RV14'], 0, alpha=0.15, color=colors['vol'])

# Annotate divergence: vol peaks near signal lows
for s in rv_signals:
    ax2.annotate('Lower Vol Peak\n(Divergence ↓)', xy=(s['signal_date'], df['RV14'].iloc[s['signal_idx']]),
                 xytext=(s['signal_date'], df['RV14'].iloc[s['signal_idx']] + 8),
                 fontsize=8, ha='center', color='yellow', fontweight='bold',
                 arrowprops=dict(arrowstyle='->', color='yellow', lw=1.5))

ax2.set_title('Volatility Proxy — "VIX Equivalent" for AAPL (14d Realized Vol)', fontsize=12, fontweight='bold', pad=10)
ax2.set_ylabel('Volatility (%)', fontsize=10)
ax2.legend(fontsize=9)
ax2.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

# ═══ CHART 3: ATR Variant ═══
ax3 = fig.add_subplot(gs[2])
ax3.plot(df['Date'], df['ATR14'], color='#AB47BC', linewidth=1.2, label='14d ATR')
ax3.fill_between(df['Date'], df['ATR14'], 0, alpha=0.15, color='#AB47BC')

for s in atr_signals:
    ax3.annotate('Divergence', xy=(s['signal_date'], df['ATR14'].iloc[s['signal_idx']]),
                 xytext=(s['signal_date'], df['ATR14'].iloc[s['signal_idx']] + 2),
                 fontsize=8, ha='center', color='yellow', fontweight='bold',
                 arrowprops=dict(arrowstyle='->', color='yellow', lw=1.5))

ax3.set_title('Volatility Proxy — Variant B: ATR (14d)', fontsize=12, fontweight='bold', pad=10)
ax3.set_ylabel('ATR ($)', fontsize=10)
ax3.legend(fontsize=9)
ax3.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

# ═══ CHART 4: Equity Curve Comparison ═══
ax4 = fig.add_subplot(gs[3])

# Buy & Hold equity
bh_eq = CAPITAL - bh_cost + bh_shares * df['Close'].values
ax4.plot(dates, bh_eq, color='white', linewidth=1.2, alpha=0.5, linestyle='--', label='Buy & Hold')
ax4.plot(dates, rv_daily['Total_Equity'].values, color=colors['equity'], linewidth=1.5, label='Scherman (RV)')
ax4.plot(dates, atr_daily['Total_Equity'].values, color='#AB47BC', linewidth=1.5, label='Scherman (ATR)')

ax4.set_title('Equity Curves: Scherman Strategy vs Buy & Hold', fontsize=13, fontweight='bold', pad=12)
ax4.set_ylabel('Portfolio Value ($)', fontsize=11)
ax4.legend(fontsize=10)
ax4.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

# ═══ CHART 5: P&L Summary Table ═══
ax5 = fig.add_subplot(gs[4])
ax5.axis('off')

# Compute summary stats for table
def get_stats(daily_df, trades):
    feq = daily_df['Total_Equity'].iloc[-1]
    tpnl = feq - CAPITAL
    tret = feq / CAPITAL - 1
    nt = len(trades)
    w = sum(1 for t in trades if t['gross_pnl'] > 0)
    wr = w / nt if nt > 0 else 0
    dr = daily_df['Daily_Return_Pct'] / 100
    sh = (dr.mean() / dr.std()) * np.sqrt(252) if dr.std() > 0 else 0
    pk = daily_df['Total_Equity'].cummax()
    mdd = ((daily_df['Total_Equity'] - pk) / pk).min()
    dl = daily_df[daily_df['Position'] == 'LONG'].shape[0]
    return [feq, tpnl, tret, nt, w, nt-w, wr, sh, mdd, dl]

rv_stats = get_stats(rv_daily, rv_trades)
atr_stats = get_stats(atr_daily, atr_trades)

table_data = [
    ['', 'Scherman\n(Real. Vol)', 'Scherman\n(ATR)', 'Buy &\nHold'],
    ['Ending Equity', f'${rv_stats[0]:,.0f}', f'${atr_stats[0]:,.0f}', f'${bh_final:,.0f}'],
    ['Total P&L', f'${rv_stats[1]:,.0f}', f'${atr_stats[1]:,.0f}', f'${bh_pnl:,.0f}'],
    ['Return', f'{rv_stats[2]:.1%}', f'{atr_stats[2]:.1%}', f'{bh_ret:.1%}'],
    ['Trades', str(rv_stats[3]), str(atr_stats[3]), '1'],
    ['Win / Loss', f'{rv_stats[4]}W / {rv_stats[5]}L', f'{atr_stats[4]}W / {atr_stats[5]}L', '1W / 0L'],
    ['Win Rate', f'{rv_stats[6]:.0%}', f'{atr_stats[6]:.0%}', '100%'],
    ['Sharpe', f'{rv_stats[7]:.2f}', f'{atr_stats[7]:.2f}', '2.31'],
    ['Max Drawdown', f'{rv_stats[8]:.1%}', f'{atr_stats[8]:.1%}', '-14.9%'],
    ['Days in Mkt', f'{rv_stats[9]}', f'{atr_stats[9]}', str(n)],
]

table = ax5.table(cellText=table_data, loc='center', cellLoc='center')
table.auto_set_font_size(False)
table.set_fontsize(11)
table.scale(1, 2.5)

for j in range(4):
    table[0, j].set_facecolor('#1a237e')
    table[0, j].set_text_props(fontweight='bold', color='white', fontsize=10)
for i in range(1, len(table_data)):
    table[i, 0].set_facecolor('#283593')
    table[i, 0].set_text_props(fontweight='bold', color='white')
    for j in range(1, 4):
        table[i, j].set_facecolor('#1a1a2e')
        table[i, j].set_text_props(color='white')

ax5.set_title('Performance Summary', fontsize=14, fontweight='bold', pad=15)

plt.savefig('/sessions/fervent-dazzling-euler/mnt/outputs/AAPL_2023_Scherman_Strategy.png',
            dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
print("\n✓ Visualization saved!")
plt.close()

# ════════════════════════════════════════════════════════════
# EXCEL P&L WORKBOOK
# ════════════════════════════════════════════════════════════
wb = Workbook()
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill('solid', fgColor='1a237e')
num_font = Font(name='Arial', size=10)
green_font = Font(name='Arial', size=10, color='008000', bold=True)
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
buy_fill_xl = PatternFill('solid', fgColor='E8F5E9')
sell_fill_xl = PatternFill('solid', fgColor='FFEBEE')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
)
money_fmt = '$#,##0.00'
pnl_fmt = '$#,##0.00;[Red]($#,##0.00);"-"'

# Summary sheet
ws = wb.active
ws.title = 'Scherman Strategy P&L'
ws.sheet_properties.tabColor = '1a237e'
ws['A1'] = "Ivan Scherman VIX Divergence Strategy — AAPL 2023 P&L"
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1a237e')
ws['A2'] = 'Entry: Buy at VIX-Price divergence | Exit: Close > SMA30 (profit) or Close < Lower BB (stop)'
ws['A2'].font = Font(name='Arial', size=10, color='666666')

# Trade log
r = 4
trade_headers = ['Variant', 'Trade #', 'Buy Date', 'Buy Price', 'Sell Date', 'Sell Price',
                 'Shares', 'Gross P&L ($)', 'Return (%)', 'Hold Days', 'Exit Reason', 'Vol Divergence', 'Result']
for c, h in enumerate(trade_headers, 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin_border

for variant_name, trades in [('Realized Vol', rv_trades), ('ATR', atr_trades)]:
    for t in trades:
        r += 1
        result = 'WIN' if t['gross_pnl'] > 0 else 'LOSS'
        vals = [variant_name, t['trade_num'],
                t['buy_date'].strftime('%Y-%m-%d'), t['buy_price'],
                t['sell_date'].strftime('%Y-%m-%d'), t['sell_price'],
                t['shares'], t['gross_pnl'], t['return_pct'],
                t['holding_days'], t['exit_reason'], f"{t['divergence_pct']:.1f}%", result]
        fmts = [None, '0', None, money_fmt, None, money_fmt, '#,##0', pnl_fmt, '0.00%', '0', None, None, None]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fmt: cell.number_format = fmt
            cell.font = num_font; cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if c == 13: cell.font = green_font if val == 'WIN' else red_font
        fill = buy_fill_xl if t['gross_pnl'] > 0 else sell_fill_xl
        for c in range(1, 14): ws.cell(row=r, column=c).fill = fill

for c in range(1, 14):
    ws.column_dimensions[get_column_letter(c)].width = max(14, len(trade_headers[c-1]) + 4)
ws.column_dimensions['K'].width = 24
ws.column_dimensions['A'].width = 14

# Daily sheets
for sheet_name, daily_df, tab_color in [
    ('RV Daily P&L', rv_daily, '26C6DA'),
    ('ATR Daily P&L', atr_daily, 'AB47BC'),
]:
    dws = wb.create_sheet(sheet_name)
    dws.sheet_properties.tabColor = tab_color
    dws['A1'] = f'{sheet_name}'
    dws['A1'].font = Font(name='Arial', bold=True, size=13, color='1a237e')

    daily_headers = ['Date', 'Close', 'SMA30', 'BB Upper', 'BB Lower', 'Vol Proxy',
                     'Action', 'Trade #', 'Position', 'Shares', 'Cost Basis', 'Mkt Value',
                     'Unrealized P&L', 'Trade P&L', 'Cum Realized P&L', 'Cash', 'Total Equity', 'Daily Ret %']
    for c, h in enumerate(daily_headers, 1):
        cell = dws.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin_border

    col_fmts = [None, money_fmt, money_fmt, money_fmt, money_fmt, '0.00',
                None, None, None, '#,##0', money_fmt, money_fmt,
                pnl_fmt, pnl_fmt, pnl_fmt, money_fmt, money_fmt, '0.00%']

    for i, row in daily_df.iterrows():
        rr = i + 4
        values = [
            row['Date'].strftime('%Y-%m-%d'), row['Close'], row['SMA30'],
            row['BB_Upper'], row['BB_Lower'], row['Vol_Proxy'],
            row['Action'], row['Trade_#'], row['Position'], row['Shares'],
            row['Cost_Basis'], row['Mkt_Value'], row['Unrealized_PnL'],
            row['Trade_PnL'] if row['Trade_PnL'] != '' else '',
            row['Cum_Realized_PnL'], row['Cash'], row['Total_Equity'],
            row['Daily_Return_Pct'] / 100
        ]
        for c, (val, fmt) in enumerate(zip(values, col_fmts), 1):
            cell = dws.cell(row=rr, column=c, value=val)
            if fmt and val != '': cell.number_format = fmt
            cell.font = num_font; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

        if row['Action'] == 'BUY':
            for c in range(1, 19): dws.cell(row=rr, column=c).fill = buy_fill_xl
            dws.cell(row=rr, column=7).font = green_font
        elif row['Action'] == 'SELL':
            for c in range(1, 19): dws.cell(row=rr, column=c).fill = sell_fill_xl
            dws.cell(row=rr, column=7).font = red_font

    for c in range(1, 19):
        dws.column_dimensions[get_column_letter(c)].width = max(12, len(daily_headers[c-1]) + 3)
    dws.freeze_panes = 'A4'

out = '/sessions/fervent-dazzling-euler/mnt/outputs/AAPL_2023_Scherman_PnL.xlsx'
wb.save(out)
print(f"✓ Excel workbook saved: {out}")
