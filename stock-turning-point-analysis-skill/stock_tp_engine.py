"""
Stock Turning Point Analysis Engine
====================================
Consolidated module for detecting price turning points using 6 mathematical
methods + ensemble consensus, backtesting, P&L computation, charting, and
Excel report generation.

Usage:
    from stock_tp_engine import run_full_analysis
    results = run_full_analysis(ticker='AAPL', period='1y', capital=100000)

    # Or with a CSV file:
    results = run_full_analysis(csv_path='AAPL_data.csv', capital=100000)

Methods implemented:
    1. Second Derivative of Gaussian-Smoothed Price
    2. Ramer-Douglas-Peucker Piecewise Linearization
    3. Wavelet Multi-Resolution Analysis (db4)
    4. PELT Change Point Detection
    5. Bry-Boschan Algorithm
    6. Menger Curvature Analysis
    7. Ivan Scherman VIX Divergence (Realized Vol + ATR variants)
    + Ensemble Consensus (Gaussian-weighted proximity voting)

Author: Quant Analysis Engine for Mohan
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
import matplotlib.gridspec as gridspec
from scipy.signal import argrelextrema
from scipy.ndimage import gaussian_filter1d
import pywt
import ruptures
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import os
from datetime import datetime

warnings.filterwarnings('ignore')


# ============================================================
# DATA LOADING
# ============================================================

def load_data(ticker=None, csv_path=None, period='1y', start_date=None, end_date=None):
    """
    Load stock price data from yfinance or a CSV file.

    Parameters:
        ticker: Stock ticker symbol (e.g., 'AAPL'). Uses yfinance.
        csv_path: Path to CSV file with Date and Close columns.
        period: yfinance period string (e.g., '1y', '2y', '6mo'). Ignored if start/end given.
        start_date: Start date string 'YYYY-MM-DD'. Overrides period.
        end_date: End date string 'YYYY-MM-DD'. Overrides period.

    Returns:
        DataFrame with Date, Close, High, Low, Open, Volume columns.
    """
    if csv_path:
        df = pd.read_csv(csv_path)
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values('Date').reset_index(drop=True)
        return df

    if ticker:
        import yfinance as yf
        if start_date and end_date:
            data = yf.download(ticker, start=start_date, end=end_date, progress=False)
        else:
            data = yf.download(ticker, period=period, progress=False)

        if data.empty:
            raise ValueError(f"No data returned for ticker '{ticker}'. Check the symbol and date range.")

        df = data.reset_index()
        # Handle multi-level columns from yfinance
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [col[0] if col[1] == '' else col[0] for col in df.columns]

        df = df.rename(columns={'Adj Close': 'Close'}) if 'Adj Close' in df.columns else df
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values('Date').reset_index(drop=True)

        required = ['Date', 'Close']
        for col in required:
            if col not in df.columns:
                raise ValueError(f"Column '{col}' not found in downloaded data. Available: {list(df.columns)}")

        return df

    raise ValueError("Must provide either 'ticker' or 'csv_path'.")


# ============================================================
# TURNING POINT DETECTION METHODS
# ============================================================

def method_second_derivative(prices, sigma=8):
    """Method 1: Second Derivative of Gaussian-Smoothed Price."""
    smoothed = gaussian_filter1d(prices, sigma=sigma)
    d1 = np.gradient(smoothed)
    d2 = np.gradient(d1)

    peaks, troughs = [], []
    for i in range(1, len(d1)):
        if d1[i-1] > 0 and d1[i] <= 0 and d2[i] < 0:
            peaks.append(i)
        elif d1[i-1] < 0 and d1[i] >= 0 and d2[i] > 0:
            troughs.append(i)

    return peaks, troughs, smoothed


def _rdp(points, epsilon):
    """Ramer-Douglas-Peucker recursive simplification."""
    if len(points) <= 2:
        return [0, len(points) - 1]
    start, end = points[0], points[-1]
    line_vec = end - start
    line_len = np.linalg.norm(line_vec)
    if line_len == 0:
        return [0, len(points) - 1]
    distances = np.abs(np.cross(line_vec, start - points)) / line_len
    max_idx = np.argmax(distances)
    if distances[max_idx] > epsilon:
        left = _rdp(points[:max_idx + 1], epsilon)
        right = [r + max_idx for r in _rdp(points[max_idx:], epsilon)]
        return left[:-1] + right
    return [0, len(points) - 1]


def method_rdp(prices, epsilon=0.02):
    """Method 2: Ramer-Douglas-Peucker Piecewise Linearization."""
    n = len(prices)
    price_norm = (prices - prices.min()) / (prices.max() - prices.min())
    points_2d = np.column_stack([np.arange(n) / n, price_norm])
    rdp_indices = _rdp(points_2d, epsilon)

    peaks, troughs = [], []
    for i, idx in enumerate(rdp_indices):
        if i == 0 or i == len(rdp_indices) - 1:
            continue
        prev_idx = rdp_indices[i - 1]
        next_idx = rdp_indices[i + 1]
        if prices[idx] > prices[prev_idx] and prices[idx] > prices[next_idx]:
            peaks.append(idx)
        elif prices[idx] < prices[prev_idx] and prices[idx] < prices[next_idx]:
            troughs.append(idx)

    return peaks, troughs, rdp_indices


def _map_to_actual(indices, prices, window=5, find_max=True):
    """Map smoothed signal indices to nearest actual price extrema."""
    mapped = []
    for idx in indices:
        s = max(0, idx - window)
        e = min(len(prices), idx + window + 1)
        if find_max:
            mapped.append(s + np.argmax(prices[s:e]))
        else:
            mapped.append(s + np.argmin(prices[s:e]))
    return mapped


def method_wavelet(prices, wavelet='db4', level=4):
    """Method 3: Wavelet Multi-Resolution Analysis."""
    n = len(prices)
    pad_len = 2 ** int(np.ceil(np.log2(n))) - n
    padded = np.pad(prices, (0, pad_len), mode='edge')
    coeffs = pywt.wavedec(padded, wavelet, level=level)
    for i in range(2, len(coeffs)):
        coeffs[i] = np.zeros_like(coeffs[i])
    reconstructed = pywt.waverec(coeffs, wavelet)[:n]

    peaks_raw = argrelextrema(reconstructed, np.greater, order=15)[0]
    troughs_raw = argrelextrema(reconstructed, np.less, order=15)[0]

    peaks = _map_to_actual(peaks_raw, prices, window=5, find_max=True)
    troughs = _map_to_actual(troughs_raw, prices, window=5, find_max=False)

    return peaks, troughs, reconstructed


def method_pelt(prices, penalty=1.5):
    """Method 4: PELT Change Point Detection."""
    n = len(prices)
    returns = np.diff(np.log(prices))
    algo = ruptures.Pelt(model="rbf", min_size=10).fit(returns.reshape(-1, 1))
    change_points = [cp for cp in algo.predict(pen=penalty) if cp < n]

    peaks, troughs = [], []
    for cp in change_points:
        if cp <= 2 or cp >= n - 2:
            continue
        lookback = min(10, cp)
        lookahead = min(10, n - cp)
        before = prices[cp] - prices[cp - lookback]
        after = prices[min(cp + lookahead, n - 1)] - prices[cp]
        if before > 0 and after < 0:
            peaks.append(cp)
        elif before < 0 and after > 0:
            troughs.append(cp)

    return peaks, troughs, change_points


def method_bry_boschan(prices, min_phase=22, min_cycle=15, window=5, min_amplitude=3.0):
    """Method 5: Bry-Boschan Algorithm (adapted for daily data)."""
    n = len(prices)
    sm = pd.Series(prices).rolling(13, center=True).mean().values
    for i in range(6):
        sm[i] = prices[i]
        sm[-(i + 1)] = prices[-(i + 1)]

    candidates = []
    for i in range(window, n - window):
        w = sm[i - window:i + window + 1]
        if sm[i] == np.nanmax(w):
            candidates.append((i, 'peak'))
        elif sm[i] == np.nanmin(w):
            candidates.append((i, 'trough'))

    if not candidates:
        return [], []

    # Enforce alternation
    alt = [candidates[0]]
    for c in candidates[1:]:
        if c[1] != alt[-1][1]:
            alt.append(c)
        elif c[1] == 'peak' and prices[c[0]] > prices[alt[-1][0]]:
            alt[-1] = c
        elif c[1] == 'trough' and prices[c[0]] < prices[alt[-1][0]]:
            alt[-1] = c

    # Enforce minimum cycle duration
    filt = [alt[0]]
    for c in alt[1:]:
        if c[0] - filt[-1][0] >= min_cycle:
            filt.append(c)
        elif c[1] == 'peak' and prices[c[0]] > prices[filt[-1][0]]:
            filt[-1] = c
        elif c[1] == 'trough' and prices[c[0]] < prices[filt[-1][0]]:
            filt[-1] = c

    # Refine on original series
    peaks_raw, troughs_raw = [], []
    for idx, tp in filt:
        s = max(0, idx - window)
        e = min(n, idx + window + 1)
        if tp == 'peak':
            peaks_raw.append(s + np.argmax(prices[s:e]))
        else:
            troughs_raw.append(s + np.argmin(prices[s:e]))

    # Enforce minimum amplitude
    all_tp = sorted(
        [(p, 'peak') for p in peaks_raw] + [(t, 'trough') for t in troughs_raw],
        key=lambda x: x[0]
    )
    peaks, troughs = [], []
    for i, (idx, tp) in enumerate(all_tp):
        if i == 0:
            (peaks if tp == 'peak' else troughs).append(idx)
            continue
        amp = abs(prices[idx] - prices[all_tp[i - 1][0]]) / prices[all_tp[i - 1][0]] * 100
        if amp >= min_amplitude:
            (peaks if tp == 'peak' else troughs).append(idx)

    return peaks, troughs


def method_menger_curvature(prices, smoothed, step=5, percentile=85):
    """Method 6: Menger Curvature Analysis."""
    n = len(prices)
    t_norm = np.arange(n) / n
    p_norm = (smoothed - smoothed.min()) / (smoothed.max() - smoothed.min())

    kappa = np.zeros(n)
    for i in range(step, n - step):
        x1, y1 = t_norm[i - step], p_norm[i - step]
        x2, y2 = t_norm[i], p_norm[i]
        x3, y3 = t_norm[i + step], p_norm[i + step]
        area = abs((x2 - x1) * (y3 - y1) - (x3 - x1) * (y2 - y1)) / 2
        d12 = np.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
        d23 = np.sqrt((x3 - x2) ** 2 + (y3 - y2) ** 2)
        d13 = np.sqrt((x3 - x1) ** 2 + (y3 - y1) ** 2)
        denom = d12 * d23 * d13
        if denom > 0:
            kappa[i] = 4 * area / denom

    thresh = np.percentile(kappa[kappa > 0], percentile) if np.any(kappa > 0) else 0
    hc = np.where(kappa > thresh)[0]

    peaks, troughs = [], []
    if len(hc) > 0:
        clusters, cur = [], [hc[0]]
        for i in range(1, len(hc)):
            if hc[i] - hc[i - 1] <= 5:
                cur.append(hc[i])
            else:
                clusters.append(cur)
                cur = [hc[i]]
        clusters.append(cur)

        dd = np.gradient(smoothed)
        for cl in clusters:
            mi = cl[np.argmax(kappa[cl])]
            if 0 < mi < n - 1:
                if dd[mi - 1] > 0 and dd[min(mi + 1, n - 1)] < 0:
                    peaks.append(mi)
                elif dd[mi - 1] < 0 and dd[min(mi + 1, n - 1)] > 0:
                    troughs.append(mi)

    return peaks, troughs, kappa


# ============================================================
# METHOD 7: IVAN SCHERMAN VIX DIVERGENCE
# ============================================================

def _find_swing_lows(series, order=5):
    """Find local minima (swing lows) in a series."""
    lows = []
    vals = series.values if hasattr(series, 'values') else np.array(series)
    for i in range(order, len(vals) - order):
        if np.isnan(vals[i]):
            continue
        window = vals[i - order:i + order + 1]
        if not np.any(np.isnan(window)) and vals[i] == np.min(window):
            lows.append(i)
    return lows


def _find_swing_highs(series, order=5):
    """Find local maxima (swing highs) in a series."""
    highs = []
    vals = series.values if hasattr(series, 'values') else np.array(series)
    for i in range(order, len(vals) - order):
        if np.isnan(vals[i]):
            continue
        window = vals[i - order:i + order + 1]
        if not np.any(np.isnan(window)) and vals[i] == np.max(window):
            highs.append(i)
    return highs


def method_scherman(df, vol_col='RV14', order=5, lookback=20):
    """
    Method 7: Ivan Scherman VIX Divergence Strategy.

    Detects buy signals where:
      1. Price makes a LOWER LOW (swing low < previous swing low)
      2. Volatility proxy makes a LOWER HIGH (divergence — fear decreasing)
      3. Price is within Bollinger Bands (30-day, 2 sigma)

    Parameters:
        df: DataFrame with Close, SMA30, BB_lower, BB_upper, and vol_col columns
        vol_col: Column name for the volatility proxy ('RV14' or 'ATR14')
        order: Window for swing detection
        lookback: Days to search for matching vol peaks near price lows

    Returns:
        signals: list of signal dicts with signal_idx, dates, prices, divergence info
        troughs: list of signal indices (for compatibility with other methods)
    """
    price_lows = _find_swing_lows(df['Close'], order=order)
    vol_highs = _find_swing_highs(df[vol_col], order=order)

    signals = []
    troughs = []

    for i in range(1, len(price_lows)):
        curr_low_idx = price_lows[i]
        prev_low_idx = price_lows[i - 1]

        # Condition 1: Price makes a LOWER LOW
        if df['Close'].iloc[curr_low_idx] >= df['Close'].iloc[prev_low_idx]:
            continue

        # Condition 2: Find vol highs near each price low
        prev_vol_highs = [vh for vh in vol_highs
                          if prev_low_idx - lookback <= vh <= prev_low_idx + lookback]
        curr_vol_highs = [vh for vh in vol_highs
                          if curr_low_idx - lookback <= vh <= curr_low_idx + lookback]

        if not prev_vol_highs or not curr_vol_highs:
            continue

        prev_vol_peak = max(prev_vol_highs, key=lambda x: df[vol_col].iloc[x])
        curr_vol_peak = max(curr_vol_highs, key=lambda x: df[vol_col].iloc[x])

        # DIVERGENCE: vol high near current low is LOWER than near previous low
        if df[vol_col].iloc[curr_vol_peak] >= df[vol_col].iloc[prev_vol_peak]:
            continue

        # Condition 3: Price within Bollinger Bands
        if pd.isna(df['BB_lower'].iloc[curr_low_idx]) or pd.isna(df['BB_upper'].iloc[curr_low_idx]):
            continue
        if df['Close'].iloc[curr_low_idx] < df['BB_lower'].iloc[curr_low_idx]:
            continue

        # All conditions met
        signals.append({
            'signal_idx': curr_low_idx,
            'signal_date': df['Date'].iloc[curr_low_idx],
            'signal_price': df['Close'].iloc[curr_low_idx],
            'prev_low_idx': prev_low_idx,
            'prev_low_price': df['Close'].iloc[prev_low_idx],
            'prev_vol_peak': df[vol_col].iloc[prev_vol_peak],
            'curr_vol_peak': df[vol_col].iloc[curr_vol_peak],
            'divergence_pct': ((df[vol_col].iloc[curr_vol_peak] / df[vol_col].iloc[prev_vol_peak]) - 1) * 100,
        })
        troughs.append(curr_low_idx)

    return signals, troughs


def compute_scherman_indicators(df):
    """
    Compute all indicators needed for the Scherman strategy.
    Adds SMA30, Bollinger Bands, Realized Volatility, and ATR to the DataFrame.
    Returns the modified DataFrame.
    """
    df = df.copy()

    # 30-day SMA
    df['SMA30'] = df['Close'].rolling(30).mean()

    # Bollinger Bands (30-day, 2 sigma)
    df['BB_std'] = df['Close'].rolling(30).std()
    df['BB_upper'] = df['SMA30'] + 2 * df['BB_std']
    df['BB_lower'] = df['SMA30'] - 2 * df['BB_std']

    # Log returns
    df['Log_Return'] = np.log(df['Close'] / df['Close'].shift(1))

    # Realized Volatility (14-day, annualized)
    df['RV14'] = df['Log_Return'].rolling(14).std() * np.sqrt(252) * 100

    # ATR (14-day) — requires High/Low columns
    if 'High' in df.columns and 'Low' in df.columns:
        df['TR'] = np.maximum(
            df['High'] - df['Low'],
            np.maximum(abs(df['High'] - df['Close'].shift(1)),
                       abs(df['Low'] - df['Close'].shift(1)))
        )
        df['ATR14'] = df['TR'].rolling(14).mean()
    else:
        # Fallback: estimate ATR from close-to-close range
        df['ATR14'] = df['Close'].rolling(14).apply(lambda x: x.max() - x.min(), raw=True)

    return df


def execute_scherman_trades(df, signals, capital=100000.0):
    """
    Execute Scherman strategy trades with specific exit rules:
      - PROFIT EXIT: Close > 30-day SMA
      - STOP LOSS: Close < lower Bollinger Band

    Returns:
        daily_df: DataFrame with daily P&L tracking
        completed_trades: list of completed trade dicts
    """
    n = len(df)
    cash = capital
    shares = 0
    cost_basis = 0
    position = 'CASH'
    trade_num = 0
    cum_realized = 0.0
    current_signal = None
    trades = []
    daily_rows = []

    signal_indices = set(s['signal_idx'] for s in signals)

    for i in range(n):
        action = ''
        trade_pnl = 0.0
        close = df['Close'].iloc[i]
        sma30 = df['SMA30'].iloc[i]
        bb_lower = df['BB_lower'].iloc[i]

        # Entry
        if position == 'CASH' and i in signal_indices:
            action = 'BUY'
            trade_num += 1
            shares = int(cash / close)
            cost_basis = shares * close
            cash -= cost_basis
            position = 'LONG'
            current_signal = next(s for s in signals if s['signal_idx'] == i)

        # Exit
        elif position == 'LONG' and not pd.isna(sma30) and not pd.isna(bb_lower):
            exit_reason = None
            if close > sma30:
                exit_reason = 'PROFIT (Close > SMA30)'
            elif close < bb_lower:
                exit_reason = 'STOP (Close < Lower BB)'

            if exit_reason:
                action = 'SELL'
                proceeds = shares * close
                trade_pnl = proceeds - cost_basis
                cum_realized += trade_pnl
                trades.append({
                    'trade_num': trade_num,
                    'buy_date': current_signal['signal_date'] if current_signal else df['Date'].iloc[i],
                    'buy_price': cost_basis / shares if shares > 0 else 0,
                    'sell_date': df['Date'].iloc[i],
                    'sell_price': close,
                    'shares': shares,
                    'gross_pnl': trade_pnl,
                    'return_pct': (proceeds / cost_basis - 1) if cost_basis > 0 else 0,
                    'exit_reason': exit_reason,
                })
                cash += proceeds
                shares = 0
                cost_basis = 0
                position = 'CASH'
                current_signal = None

        mkt_value = shares * close
        unrealized = mkt_value - cost_basis if shares > 0 else 0
        total_equity = cash + mkt_value
        prev_eq = daily_rows[-1]['Total_Equity'] if daily_rows else capital
        daily_ret = ((total_equity / prev_eq) - 1) * 100 if prev_eq > 0 else 0

        daily_rows.append({
            'Date': df['Date'].iloc[i],
            'Close': close,
            'Action': action,
            'Trade_#': trade_num if action else '',
            'Position': position,
            'Shares': shares,
            'Cost_Basis': cost_basis if shares > 0 else 0,
            'Mkt_Value': mkt_value,
            'Unrealized_PnL': unrealized,
            'Trade_PnL': trade_pnl if action == 'SELL' else '',
            'Cum_Realized_PnL': cum_realized,
            'Cash': cash,
            'Total_Equity': total_equity,
            'Daily_Return_Pct': daily_ret,
        })

    # Close open position at end
    if position == 'LONG':
        final_proceeds = shares * df['Close'].iloc[-1]
        final_pnl = final_proceeds - cost_basis
        trades.append({
            'trade_num': trade_num,
            'buy_date': current_signal['signal_date'] if current_signal else df['Date'].iloc[-1],
            'buy_price': cost_basis / shares if shares > 0 else 0,
            'sell_date': df['Date'].iloc[-1],
            'sell_price': df['Close'].iloc[-1],
            'shares': shares,
            'gross_pnl': final_pnl,
            'return_pct': (final_proceeds / cost_basis - 1) if cost_basis > 0 else 0,
            'exit_reason': 'OPEN (Period-End)',
        })

    return pd.DataFrame(daily_rows), trades


# ============================================================
# ENSEMBLE CONSENSUS
# ============================================================

def compute_ensemble(all_peaks, all_troughs, n, proximity=5, threshold=2.0):
    """
    Compute ensemble consensus turning points using Gaussian-weighted proximity voting.

    Parameters:
        all_peaks: list of (indices, weight) tuples for each method's peaks
        all_troughs: list of (indices, weight) tuples for each method's troughs
        n: total number of data points
        proximity: window size for consensus (days)
        threshold: minimum score to qualify as consensus point

    Returns:
        consensus_peaks, consensus_troughs, peak_score, trough_score
    """
    peak_score = np.zeros(n)
    trough_score = np.zeros(n)

    for score_arr, methods in [(peak_score, all_peaks), (trough_score, all_troughs)]:
        for indices, weight in methods:
            for idx in indices:
                for d in range(max(0, idx - proximity), min(n, idx + proximity + 1)):
                    dist = abs(d - idx)
                    score_arr[d] += weight * np.exp(-0.5 * (dist / 2) ** 2)

    consensus_peaks = [
        p for p in argrelextrema(peak_score, np.greater, order=10)[0]
        if peak_score[p] >= threshold
    ]
    consensus_troughs = [
        t for t in argrelextrema(trough_score, np.greater, order=10)[0]
        if trough_score[t] >= threshold
    ]

    return consensus_peaks, consensus_troughs, peak_score, trough_score


# ============================================================
# DETECTION ENGINE: RUN ALL METHODS
# ============================================================

def detect_all_turning_points(prices, df=None):
    """
    Run all 7 turning point detection methods + ensemble.

    Parameters:
        prices: numpy array of close prices
        df: DataFrame with Date, Close, High, Low columns (needed for Scherman method).
            If None, Scherman method is skipped.

    Returns:
        dict with method names as keys, each containing 'peaks' and 'troughs' lists,
        plus 'ensemble' entry with consensus results, and Scherman-specific data.
    """
    n = len(prices)
    results = {}

    # Method 1: Second Derivative
    m1_peaks, m1_troughs, smoothed = method_second_derivative(prices, sigma=8)
    results['1. 2nd Derivative'] = {'peaks': m1_peaks, 'troughs': m1_troughs}

    # Method 2: RDP
    m2_peaks, m2_troughs, rdp_idx = method_rdp(prices, epsilon=0.02)
    results['2. RDP Linearization'] = {'peaks': m2_peaks, 'troughs': m2_troughs}

    # Method 3: Wavelet
    m3_peaks, m3_troughs, reconstructed = method_wavelet(prices)
    results['3. Wavelet (db4)'] = {'peaks': m3_peaks, 'troughs': m3_troughs}

    # Method 4: PELT
    m4_peaks, m4_troughs, cps = method_pelt(prices, penalty=1.5)
    results['4. PELT Change Point'] = {'peaks': m4_peaks, 'troughs': m4_troughs}

    # Method 5: Bry-Boschan
    m5_peaks, m5_troughs = method_bry_boschan(prices)
    results['5. Bry-Boschan'] = {'peaks': m5_peaks, 'troughs': m5_troughs}

    # Method 6: Menger Curvature
    m6_peaks, m6_troughs, kappa = method_menger_curvature(prices, smoothed)
    results['6. Menger Curvature'] = {'peaks': m6_peaks, 'troughs': m6_troughs}

    # Method 7: Ivan Scherman VIX Divergence (two variants)
    scherman_df = None
    if df is not None:
        scherman_df = compute_scherman_indicators(df)

        # Variant A: Realized Volatility
        rv_signals, rv_troughs = method_scherman(scherman_df, vol_col='RV14')
        results['7a. Scherman (RealVol)'] = {
            'peaks': [], 'troughs': rv_troughs,
            'signals': rv_signals, 'type': 'scherman',
        }

        # Variant B: ATR
        atr_signals, atr_troughs = method_scherman(scherman_df, vol_col='ATR14')
        results['7b. Scherman (ATR)'] = {
            'peaks': [], 'troughs': atr_troughs,
            'signals': atr_signals, 'type': 'scherman',
        }

    # Ensemble consensus (uses methods 1-6 only, not Scherman)
    weights = {
        '1. 2nd Derivative': 1.0, '2. RDP Linearization': 0.8,
        '3. Wavelet (db4)': 1.0, '4. PELT Change Point': 0.7,
        '5. Bry-Boschan': 1.2, '6. Menger Curvature': 0.9,
    }
    all_peaks = [(results[m]['peaks'], weights[m]) for m in weights]
    all_troughs = [(results[m]['troughs'], weights[m]) for m in weights]
    ens_peaks, ens_troughs, peak_score, trough_score = compute_ensemble(
        all_peaks, all_troughs, n
    )
    results['8. Ensemble Consensus'] = {'peaks': ens_peaks, 'troughs': ens_troughs}

    # Store auxiliary data for charting
    results['_aux'] = {
        'smoothed': smoothed, 'reconstructed': reconstructed,
        'curvature': kappa, 'peak_score': peak_score, 'trough_score': trough_score,
        'scherman_df': scherman_df,
    }

    return results


# ============================================================
# TRADE ENGINE
# ============================================================

def generate_trades(peaks, troughs, prices, dates):
    """Interleave troughs (buys) and peaks (sells) chronologically with alternation."""
    events = sorted(
        [(t, 'BUY', prices[t], dates[t]) for t in troughs] +
        [(p, 'SELL', prices[p], dates[p]) for p in peaks],
        key=lambda x: x[0]
    )
    trades = []
    looking_for = 'BUY'
    for idx, action, price, date in events:
        if action == looking_for:
            trades.append({
                'idx': idx, 'action': action, 'price': price,
                'date': pd.Timestamp(date)
            })
            looking_for = 'SELL' if action == 'BUY' else 'BUY'
    return trades


def compute_pnl(trade_sequence, prices, dates, capital=100000.0):
    """
    Compute full daily P&L simulation from alternating BUY/SELL trade sequence.

    Returns:
        daily_df: DataFrame with daily P&L tracking
        completed_trades: list of completed trade dicts
    """
    n = len(prices)
    cash = capital
    shares = 0
    cost_basis = 0
    position = 'CASH'
    cum_realized = 0.0
    trade_num = 0
    completed_trades = []

    # Build position array
    positions = np.zeros(n)
    buy_idx = None
    for t in trade_sequence:
        if t['action'] == 'BUY':
            buy_idx = t['idx']
        elif t['action'] == 'SELL' and buy_idx is not None:
            positions[buy_idx:t['idx'] + 1] = 1
            buy_idx = None
    if buy_idx is not None:
        positions[buy_idx:] = 1

    daily_rows = []
    for i in range(n):
        prev_pos = positions[i - 1] if i > 0 else 0
        curr_pos = positions[i]
        action = ''
        trade_pnl = 0.0

        if curr_pos == 1 and prev_pos == 0:
            action = 'BUY'
            trade_num += 1
            shares = int(cash / prices[i])
            cost_basis = shares * prices[i]
            cash -= cost_basis
            position = 'LONG'
        elif curr_pos == 0 and prev_pos == 1:
            action = 'SELL'
            proceeds = shares * prices[i]
            trade_pnl = proceeds - cost_basis
            cum_realized += trade_pnl
            completed_trades.append({
                'trade_num': trade_num,
                'buy_date': None,
                'sell_date': pd.Timestamp(dates[i]),
                'buy_price': cost_basis / shares if shares > 0 else 0,
                'sell_price': prices[i],
                'shares': shares,
                'gross_pnl': trade_pnl,
                'return_pct': (proceeds / cost_basis - 1) if cost_basis > 0 else 0,
            })
            cash += proceeds
            shares = 0
            cost_basis = 0
            position = 'CASH'

        mkt_value = shares * prices[i]
        unrealized = mkt_value - cost_basis if shares > 0 else 0
        total_equity = cash + mkt_value
        prev_equity = daily_rows[-1]['Total_Equity'] if daily_rows else capital
        daily_ret = ((total_equity / prev_equity) - 1) * 100 if prev_equity > 0 else 0

        daily_rows.append({
            'Date': pd.Timestamp(dates[i]),
            'Close': prices[i],
            'Action': action,
            'Trade_#': trade_num if action else '',
            'Position': position,
            'Shares': shares,
            'Cost_Basis': cost_basis if shares > 0 else 0,
            'Mkt_Value': mkt_value,
            'Unrealized_PnL': unrealized,
            'Trade_PnL': trade_pnl if action == 'SELL' else '',
            'Cum_Realized_PnL': cum_realized,
            'Cash': cash,
            'Total_Equity': total_equity,
            'Daily_Return_Pct': daily_ret,
        })

    # Fill buy_dates in completed trades
    buy_dates = [t for t in trade_sequence if t['action'] == 'BUY']
    for i, ct in enumerate(completed_trades):
        if i < len(buy_dates):
            ct['buy_date'] = buy_dates[i]['date']

    return pd.DataFrame(daily_rows), completed_trades


def compute_buy_hold(prices, dates, capital=100000.0):
    """Compute Buy & Hold P&L for benchmark comparison."""
    n = len(prices)
    shares = int(capital / prices[0])
    cost = shares * prices[0]
    cash = capital - cost
    rows = []
    for i in range(n):
        mv = shares * prices[i]
        eq = cash + mv
        prev_eq = rows[-1]['Total_Equity'] if rows else capital
        rows.append({
            'Date': pd.Timestamp(dates[i]), 'Close': prices[i],
            'Action': 'BUY' if i == 0 else ('SELL' if i == n - 1 else ''),
            'Trade_#': 1 if i == 0 or i == n - 1 else '',
            'Position': 'LONG', 'Shares': shares,
            'Cost_Basis': cost, 'Mkt_Value': mv,
            'Unrealized_PnL': mv - cost,
            'Trade_PnL': mv - cost if i == n - 1 else '',
            'Cum_Realized_PnL': 0 if i < n - 1 else mv - cost,
            'Cash': cash, 'Total_Equity': eq,
            'Daily_Return_Pct': ((eq / prev_eq) - 1) * 100 if prev_eq > 0 else 0,
        })
    bh_df = pd.DataFrame(rows)
    bh_completed = [{
        'trade_num': 1, 'buy_date': pd.Timestamp(dates[0]),
        'sell_date': pd.Timestamp(dates[-1]),
        'buy_price': prices[0], 'sell_price': prices[-1], 'shares': shares,
        'gross_pnl': shares * (prices[-1] - prices[0]),
        'return_pct': prices[-1] / prices[0] - 1,
    }]
    return bh_df, bh_completed


def compute_summary_stats(daily_df, completed_trades, capital, n):
    """Compute summary statistics for a strategy."""
    final_eq = daily_df['Total_Equity'].iloc[-1]
    total_pnl = final_eq - capital
    total_ret = final_eq / capital - 1
    num_trades = len(completed_trades)
    winners = sum(1 for t in completed_trades if t['gross_pnl'] > 0)
    losers = num_trades - winners
    win_rate = winners / num_trades if num_trades > 0 else 0
    daily_rets = daily_df['Daily_Return_Pct'] / 100
    vol = daily_rets.std() * np.sqrt(252)
    sharpe = (daily_rets.mean() / daily_rets.std()) * np.sqrt(252) if daily_rets.std() > 0 else 0
    peak = daily_df['Total_Equity'].cummax()
    max_dd = ((daily_df['Total_Equity'] - peak) / peak).min()
    days_long = daily_df[daily_df['Position'] == 'LONG'].shape[0]
    exposure = days_long / n
    best_t = max([t['return_pct'] for t in completed_trades]) if completed_trades else 0
    worst_t = min([t['return_pct'] for t in completed_trades]) if completed_trades else 0
    avg_t = np.mean([t['return_pct'] for t in completed_trades]) if completed_trades else 0

    return {
        'final_eq': final_eq, 'total_pnl': total_pnl, 'total_ret': total_ret,
        'num_trades': num_trades, 'winners': winners, 'losers': losers,
        'win_rate': win_rate, 'best_trade': best_t, 'worst_trade': worst_t,
        'avg_trade': avg_t, 'vol': vol, 'sharpe': sharpe, 'max_dd': max_dd,
        'exposure': exposure, 'days_long': days_long,
    }


# ============================================================
# BACKTEST ALL STRATEGIES
# ============================================================

def backtest_all(detection_results, prices, dates, capital=100000.0):
    """
    Run P&L backtests for all detection methods + ensemble + Buy & Hold.

    Returns:
        strategies: dict of {name: (daily_df, completed_trades, summary_stats)}
    """
    n = len(prices)
    strategies = {}

    # Buy & Hold
    bh_df, bh_completed = compute_buy_hold(prices, dates, capital)
    bh_stats = compute_summary_stats(bh_df, bh_completed, capital, n)
    bh_stats['name'] = 'Buy & Hold'
    strategies['Buy & Hold'] = (bh_df, bh_completed, bh_stats)

    # Each detection method
    scherman_df = detection_results.get('_aux', {}).get('scherman_df')
    method_names = [k for k in detection_results if not k.startswith('_')]
    for name in method_names:
        method_data = detection_results[name]

        # Scherman methods use their own trade execution engine
        if method_data.get('type') == 'scherman' and scherman_df is not None:
            signals = method_data.get('signals', [])
            daily_df, completed = execute_scherman_trades(scherman_df, signals, capital)
            stats = compute_summary_stats(daily_df, completed, capital, n)
            stats['name'] = name
            strategies[name] = (daily_df, completed, stats)
        else:
            peaks = method_data['peaks']
            troughs = method_data['troughs']
            seq = generate_trades(peaks, troughs, prices, dates)
            daily_df, completed = compute_pnl(seq, prices, dates, capital)
            stats = compute_summary_stats(daily_df, completed, capital, n)
            stats['name'] = name
            strategies[name] = (daily_df, completed, stats)

    return strategies


# ============================================================
# METHOD EVALUATION & RANKING
# ============================================================

def evaluate_methods(detection_results, prices, dates, proximity=5):
    """Evaluate and rank each method against ensemble consensus."""
    methods = [k for k in detection_results
               if not k.startswith('_') and k != '8. Ensemble Consensus'
               and not detection_results[k].get('type') == 'scherman']
    ens = detection_results.get('8. Ensemble Consensus', {'peaks': [], 'troughs': []})
    consensus_peaks = ens['peaks']
    consensus_troughs = ens['troughs']

    rankings = []
    for name in methods:
        peaks = detection_results[name]['peaks']
        troughs = detection_results[name]['troughs']
        all_detected = list(peaks) + list(troughs)
        total = len(all_detected)
        if total == 0:
            rankings.append({'name': name, 'total': 0, 'consensus_match': 0,
                             'precision': 0, 'avg_amplitude': 0, 'score': 0})
            continue

        matches = 0
        for p in peaks:
            for cp in consensus_peaks:
                if abs(p - cp) <= proximity:
                    matches += 1
                    break
        for t in troughs:
            for ct in consensus_troughs:
                if abs(t - ct) <= proximity:
                    matches += 1
                    break

        precision = matches / total
        amplitudes = []
        for idx in all_detected:
            s = max(0, idx - 10)
            e = min(len(prices), idx + 10)
            local_range = (prices[s:e].max() - prices[s:e].min()) / prices[idx] * 100
            amplitudes.append(local_range)
        avg_amp = np.mean(amplitudes)

        score = precision * 60 + min(avg_amp, 15) * 2 + (10 if 3 <= total <= 8 else 0)
        rankings.append({
            'name': name, 'total': total, 'consensus_match': matches,
            'precision': precision * 100, 'avg_amplitude': avg_amp, 'score': score,
        })

    rankings.sort(key=lambda x: x['score'], reverse=True)
    return rankings


# ============================================================
# CHART GENERATION
# ============================================================

def generate_charts(df, detection_results, strategies, output_dir, ticker_label):
    """Generate all visualization charts."""
    prices = df['Close'].values
    dates = df['Date'].values
    n = len(prices)
    aux = detection_results.get('_aux', {})
    smoothed = aux.get('smoothed', prices)
    reconstructed = aux.get('reconstructed', prices)
    kappa = aux.get('curvature', np.zeros(n))
    peak_score = aux.get('peak_score', np.zeros(n))
    trough_score = aux.get('trough_score', np.zeros(n))

    ens = detection_results.get('8. Ensemble Consensus', {'peaks': [], 'troughs': []})
    consensus_peaks = ens['peaks']
    consensus_troughs = ens['troughs']

    colors = {
        'price': '#42A5F5', 'peak': '#FF1744', 'trough': '#00E676',
        'smooth': '#FFC107',
    }

    # ---- MAIN CHART: Turning Points ----
    fig = plt.figure(figsize=(22, 32))
    gs = gridspec.GridSpec(7, 2, height_ratios=[3, 2, 2, 2, 2, 2, 3], hspace=0.35, wspace=0.25)

    ax_main = fig.add_subplot(gs[0, :])
    ax_main.plot(dates, prices, color=colors['price'], linewidth=1.2, alpha=0.8, label=f'{ticker_label} Close')

    for p in consensus_peaks:
        ax_main.annotate(
            f'${prices[p]:.0f}\n{pd.Timestamp(dates[p]).strftime("%b %d")}',
            xy=(dates[p], prices[p]), xytext=(dates[p], prices[p] + (prices.max() - prices.min()) * 0.03),
            fontsize=8, ha='center', color=colors['peak'], fontweight='bold',
            arrowprops=dict(arrowstyle='->', color=colors['peak'], lw=1.5))
        ax_main.scatter([dates[p]], [prices[p]], color=colors['peak'], s=200, zorder=5,
                        marker='v', edgecolors='white', linewidth=1)

    for t in consensus_troughs:
        ax_main.annotate(
            f'${prices[t]:.0f}\n{pd.Timestamp(dates[t]).strftime("%b %d")}',
            xy=(dates[t], prices[t]), xytext=(dates[t], prices[t] - (prices.max() - prices.min()) * 0.04),
            fontsize=8, ha='center', color=colors['trough'], fontweight='bold',
            arrowprops=dict(arrowstyle='->', color=colors['trough'], lw=1.5))
        ax_main.scatter([dates[t]], [prices[t]], color=colors['trough'], s=200, zorder=5,
                        marker='^', edgecolors='white', linewidth=1)

    ax_main.set_title(f'{ticker_label} — Ensemble Consensus Turning Points', fontsize=16, fontweight='bold', pad=15)
    ax_main.set_ylabel('Price ($)', fontsize=12)
    ax_main.legend(fontsize=10, loc='upper left')
    ax_main.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
    ax_main.xaxis.set_major_locator(mdates.MonthLocator())

    # Individual method charts
    # Methods 1-6 for the individual sub-charts (Scherman and Ensemble excluded)
    method_keys = [k for k in detection_results
                   if not k.startswith('_') and k != '8. Ensemble Consensus'
                   and not detection_results[k].get('type') == 'scherman']
    overlays = {
        '1. 2nd Derivative': smoothed,
        '3. Wavelet (db4)': reconstructed,
    }

    for i, key in enumerate(method_keys):
        row = 1 + i // 2
        col = i % 2
        ax = fig.add_subplot(gs[row, col])
        ax.plot(dates, prices, color=colors['price'], linewidth=1, alpha=0.6)
        if key in overlays:
            ax.plot(dates, overlays[key], color=colors['smooth'], linewidth=1.2, alpha=0.8, linestyle='--')

        peaks = detection_results[key]['peaks']
        troughs = detection_results[key]['troughs']
        if peaks:
            ax.scatter([dates[p] for p in peaks], [prices[p] for p in peaks],
                       color=colors['peak'], s=100, marker='v', zorder=5, edgecolors='white', linewidth=0.5)
        if troughs:
            ax.scatter([dates[t] for t in troughs], [prices[t] for t in troughs],
                       color=colors['trough'], s=100, marker='^', zorder=5, edgecolors='white', linewidth=0.5)

        all_tp = sorted([(p, 'peak') for p in peaks] + [(t, 'trough') for t in troughs], key=lambda x: x[0])
        if len(all_tp) > 1:
            tp_dates = [dates[tp[0]] for tp in all_tp]
            tp_prices = [prices[tp[0]] for tp in all_tp]
            ax.plot(tp_dates, tp_prices, color='white', linewidth=0.8, alpha=0.4, linestyle='-')

        ax.set_title(f'Method {key}', fontsize=11, fontweight='bold', pad=8)
        ax.set_ylabel('$', fontsize=9)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        ax.tick_params(labelsize=8)

    # Consensus score chart
    ax_score = fig.add_subplot(gs[4, :])
    ax_score.fill_between(dates, peak_score, 0, alpha=0.4, color=colors['peak'], label='Peak Score')
    ax_score.fill_between(dates, -trough_score, 0, alpha=0.4, color=colors['trough'], label='Trough Score')
    ax_score.axhline(y=2.0, color='white', linestyle='--', linewidth=0.8, alpha=0.5, label='Threshold (2.0)')
    ax_score.axhline(y=-2.0, color='white', linestyle='--', linewidth=0.8, alpha=0.5)
    ax_score.set_title('Ensemble Consensus Score', fontsize=13, fontweight='bold', pad=10)
    ax_score.set_ylabel('Score', fontsize=11)
    ax_score.legend(fontsize=9, loc='upper right')
    ax_score.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

    # Curvature chart
    ax_curv = fig.add_subplot(gs[5, 0])
    ax_curv.plot(dates, kappa, color='#AB47BC', linewidth=0.8)
    curv_thresh = np.percentile(kappa[kappa > 0], 85) if np.any(kappa > 0) else 0
    ax_curv.axhline(y=curv_thresh, color='yellow', linestyle='--', linewidth=0.8, alpha=0.7)
    ax_curv.fill_between(dates, kappa, 0, where=kappa > curv_thresh, alpha=0.3, color='yellow')
    ax_curv.set_title('Menger Curvature', fontsize=11, fontweight='bold', pad=8)
    ax_curv.set_ylabel('kappa', fontsize=11)
    ax_curv.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

    # Second derivative chart
    d2 = np.gradient(np.gradient(smoothed))
    ax_d2 = fig.add_subplot(gs[5, 1])
    ax_d2.plot(dates, d2, color='#26C6DA', linewidth=0.8)
    ax_d2.fill_between(dates, d2, 0, where=np.array(d2) > 0, alpha=0.3, color='#00E676')
    ax_d2.fill_between(dates, d2, 0, where=np.array(d2) < 0, alpha=0.3, color='#FF1744')
    ax_d2.axhline(y=0, color='white', linewidth=0.5, alpha=0.5)
    ax_d2.set_title('2nd Derivative (Concavity)', fontsize=11, fontweight='bold', pad=8)
    ax_d2.set_ylabel("f''(x)", fontsize=11)
    ax_d2.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

    # Ranking table
    rankings = evaluate_methods(detection_results, prices, dates)
    ax_table = fig.add_subplot(gs[6, :])
    ax_table.axis('off')
    table_data = [['Rank', 'Method', 'Detections', 'Consensus\nMatches', 'Precision', 'Avg\nAmplitude', 'Score']]
    for i, r in enumerate(rankings):
        table_data.append([
            f'#{i+1}', r['name'], str(r['total']), str(r['consensus_match']),
            f"{r['precision']:.0f}%", f"{r['avg_amplitude']:.1f}%", f"{r['score']:.1f}"
        ])
    table = ax_table.table(cellText=table_data, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1, 2.5)
    for j in range(7):
        table[0, j].set_facecolor('#1a237e')
        table[0, j].set_text_props(fontweight='bold', color='white', fontsize=10)
    for i in range(1, len(table_data)):
        for j in range(7):
            table[i, j].set_facecolor('#1a1a2e')
            table[i, j].set_text_props(color='white')
        if i == 1:
            for j in range(7):
                table[i, j].set_facecolor('#1B5E20')
                table[i, j].set_text_props(fontweight='bold', color='#00E676')
    ax_table.set_title('Method Ranking by Composite Score', fontsize=14, fontweight='bold', pad=15)

    tp_chart_path = os.path.join(output_dir, f'{ticker_label}_Turning_Points.png')
    plt.savefig(tp_chart_path, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    print(f"  Chart saved: {tp_chart_path}")

    # ---- CUMULATIVE RETURNS CHART ----
    fig2, ax = plt.subplots(figsize=(18, 8))
    strategy_colors = ['#3F51B5', '#4CAF50', '#2196F3', '#AB47BC', '#FF9800', '#009688', '#E91E63', '#FF5722']
    all_names = list(strategies.keys())
    for i, name in enumerate(all_names):
        daily_df = strategies[name][0]
        equity = daily_df['Total_Equity'].values
        cum_ret = (equity / equity[0] - 1) * 100
        style = '--' if name == 'Buy & Hold' else '-'
        ax.plot(daily_df['Date'], cum_ret, linewidth=1.5 if name != 'Buy & Hold' else 2,
                linestyle=style, label=f"{name} ({cum_ret[-1]:+.1f}%)",
                color=strategy_colors[i % len(strategy_colors)])

    ax.set_title(f'{ticker_label} — Cumulative Returns: All Strategies', fontsize=14, fontweight='bold')
    ax.set_ylabel('Cumulative Return (%)', fontsize=11)
    ax.legend(fontsize=8, loc='upper left', ncol=2)
    ax.axhline(y=0, color='white', linewidth=0.5, alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b'))

    ret_chart_path = os.path.join(output_dir, f'{ticker_label}_Cumulative_Returns.png')
    plt.savefig(ret_chart_path, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    print(f"  Chart saved: {ret_chart_path}")

    return tp_chart_path, ret_chart_path


# ============================================================
# EXCEL WORKBOOK GENERATION
# ============================================================

def generate_excel(strategies, output_dir, ticker_label, capital=100000.0):
    """Generate professional Excel P&L workbook."""
    wb = Workbook()
    n = max(len(strategies[k][0]) for k in strategies)

    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='1a237e')
    title_font = Font(name='Arial', bold=True, size=14, color='1a237e')
    num_font = Font(name='Arial', size=10)
    green_font = Font(name='Arial', size=10, color='008000', bold=True)
    red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
    buy_fill = PatternFill('solid', fgColor='E8F5E9')
    sell_fill = PatternFill('solid', fgColor='FFEBEE')
    best_fill = PatternFill('solid', fgColor='C8E6C9')
    total_fill = PatternFill('solid', fgColor='E3F2FD')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    money_fmt = '$#,##0.00'
    pnl_fmt = '$#,##0.00;[Red]($#,##0.00);"-"'

    # ---- SUMMARY SHEET ----
    ws = wb.active
    ws.title = 'P&L Summary'
    ws.sheet_properties.tabColor = '1a237e'
    ws['A1'] = f'{ticker_label} — Turning Point Strategy P&L Comparison'
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='1a237e')
    ws['A2'] = f'Starting Capital: ${capital:,.0f} | Logic: BUY at detected troughs, SELL at detected peaks'
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    all_names = list(strategies.keys())
    headers = ['Metric'] + all_names
    r = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    metrics_order = [
        ('Starting Capital', lambda s: capital, money_fmt),
        ('Ending Equity', lambda s: s['final_eq'], money_fmt),
        ('Total P&L ($)', lambda s: s['total_pnl'], pnl_fmt),
        ('Total Return (%)', lambda s: s['total_ret'], '0.00%'),
        ('Number of Trades', lambda s: s['num_trades'], '0'),
        ('Winning Trades', lambda s: s['winners'], '0'),
        ('Losing Trades', lambda s: s['losers'], '0'),
        ('Win Rate (%)', lambda s: s['win_rate'], '0.0%'),
        ('Best Trade (%)', lambda s: s['best_trade'], '0.00%'),
        ('Worst Trade (%)', lambda s: s['worst_trade'], '0.00%'),
        ('Avg Trade Return (%)', lambda s: s['avg_trade'], '0.00%'),
        ('Annualized Volatility (%)', lambda s: s['vol'], '0.00%'),
        ('Sharpe Ratio', lambda s: s['sharpe'], '0.00'),
        ('Max Drawdown (%)', lambda s: s['max_dd'], '0.00%'),
        ('Market Exposure (%)', lambda s: s['exposure'], '0.0%'),
        ('Days in Market', lambda s: s['days_long'], '0'),
        ('Days in Cash', lambda s: n - s['days_long'], '0'),
    ]

    for mi, (metric_name, func, fmt) in enumerate(metrics_order):
        row = r + 1 + mi
        cell = ws.cell(row=row, column=1, value=metric_name)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = thin_border

        vals = []
        for ci, name in enumerate(all_names):
            stats = strategies[name][2]
            val = func(stats)
            vals.append(val)
            cell = ws.cell(row=row, column=ci + 2, value=val)
            cell.number_format = fmt
            cell.font = num_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if metric_name in ['Total P&L ($)', 'Total Return (%)']:
                cell.font = green_font if val > 0 else (red_font if val < 0 else num_font)
            if metric_name == 'Ending Equity':
                cell.fill = total_fill

    ws.column_dimensions['A'].width = 28
    for c in range(2, len(all_names) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ---- TRADE LOG SHEET ----
    tl = wb.create_sheet('Trade Log')
    tl.sheet_properties.tabColor = 'FF7043'
    tl['A1'] = 'Trade-by-Trade P&L — All Strategies'
    tl['A1'].font = title_font

    trade_headers = ['Strategy', 'Trade #', 'Buy Date', 'Buy Price', 'Sell Date', 'Sell Price',
                     'Shares', 'Gross P&L ($)', 'Return (%)', 'Holding Days', 'Result']
    tr = 3
    for c, h in enumerate(trade_headers, 1):
        cell = tl.cell(row=tr, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for name in all_names:
        completed = strategies[name][1]
        for t in completed:
            tr += 1
            days = (t['sell_date'] - t['buy_date']).days if t['buy_date'] else 0
            result = 'WIN' if t['gross_pnl'] > 0 else 'LOSS'
            vals = [
                name, t['trade_num'],
                t['buy_date'].strftime('%Y-%m-%d') if t['buy_date'] else '',
                t['buy_price'], t['sell_date'].strftime('%Y-%m-%d'),
                t['sell_price'], t['shares'], t['gross_pnl'], t['return_pct'], days, result
            ]
            fmts = [None, '0', None, money_fmt, None, money_fmt, '#,##0', pnl_fmt, '0.00%', '0', None]
            for c, (val, fmt_str) in enumerate(zip(vals, fmts), 1):
                cell = tl.cell(row=tr, column=c, value=val)
                if fmt_str:
                    cell.number_format = fmt_str
                cell.font = num_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if c == 11:
                    cell.font = green_font if val == 'WIN' else red_font
            fill = buy_fill if t['gross_pnl'] > 0 else sell_fill
            for c in range(1, 12):
                tl.cell(row=tr, column=c).fill = fill

    for c in range(1, 12):
        tl.column_dimensions[get_column_letter(c)].width = max(14, len(trade_headers[c - 1]) + 4)
    tl.column_dimensions['A'].width = 24
    tl.freeze_panes = 'A4'

    # ---- DAILY P&L SHEETS ----
    tab_colors = ['4CAF50', '2196F3', 'AB47BC', 'FF9800', '009688', 'E91E63', 'FF5722', '3F51B5']
    daily_headers = ['Date', 'Close', 'Action', 'Trade #', 'Position', 'Shares',
                     'Cost Basis', 'Mkt Value', 'Unrealized P&L', 'Trade P&L',
                     'Cum Realized P&L', 'Cash', 'Total Equity', 'Daily Return %']
    col_fmts = [None, money_fmt, None, None, None, '#,##0',
                money_fmt, money_fmt, pnl_fmt, pnl_fmt, pnl_fmt, money_fmt, money_fmt, '0.00%']

    for si, name in enumerate(all_names):
        daily_df = strategies[name][0]
        short_name = name[:25]
        ws2 = wb.create_sheet(short_name)
        ws2.sheet_properties.tabColor = tab_colors[si % len(tab_colors)]
        ws2['A1'] = f'{name} — Daily P&L'
        ws2['A1'].font = title_font

        for c, h in enumerate(daily_headers, 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for i, row_data in daily_df.iterrows():
            rr = i + 4
            values = [
                row_data['Date'].strftime('%Y-%m-%d'), row_data['Close'], row_data['Action'],
                row_data['Trade_#'], row_data['Position'], row_data['Shares'],
                row_data['Cost_Basis'], row_data['Mkt_Value'], row_data['Unrealized_PnL'],
                row_data['Trade_PnL'] if row_data['Trade_PnL'] != '' else '',
                row_data['Cum_Realized_PnL'], row_data['Cash'], row_data['Total_Equity'],
                row_data['Daily_Return_Pct'] / 100,
            ]
            for c, (val, fmt_str) in enumerate(zip(values, col_fmts), 1):
                cell = ws2.cell(row=rr, column=c, value=val)
                if fmt_str and val != '':
                    cell.number_format = fmt_str
                cell.font = num_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            if row_data['Action'] == 'BUY':
                for c in range(1, 15):
                    ws2.cell(row=rr, column=c).fill = buy_fill
                ws2.cell(row=rr, column=3).font = green_font
            elif row_data['Action'] == 'SELL':
                for c in range(1, 15):
                    ws2.cell(row=rr, column=c).fill = sell_fill
                ws2.cell(row=rr, column=3).font = red_font

        for c in range(1, 15):
            ws2.column_dimensions[get_column_letter(c)].width = max(13, len(daily_headers[c - 1]) + 3)
        ws2.freeze_panes = 'A4'

    xlsx_path = os.path.join(output_dir, f'{ticker_label}_TurningPoint_PnL.xlsx')
    wb.save(xlsx_path)
    print(f"  Excel saved: {xlsx_path}")
    return xlsx_path


# ============================================================
# CSV EXPORT
# ============================================================

def generate_csv(strategies, rankings, output_dir, ticker_label):
    """Export strategy summary and trade log to CSV."""
    # Strategy summary
    summary_rows = []
    for name in strategies:
        stats = strategies[name][2]
        summary_rows.append({
            'Strategy': name,
            'Final Equity': f"${stats['final_eq']:,.2f}",
            'Total P&L': f"${stats['total_pnl']:,.2f}",
            'Return': f"{stats['total_ret']:.2%}",
            'Trades': stats['num_trades'],
            'Win Rate': f"{stats['win_rate']:.0%}",
            'Sharpe': f"{stats['sharpe']:.2f}",
            'Max Drawdown': f"{stats['max_dd']:.2%}",
        })
    summary_path = os.path.join(output_dir, f'{ticker_label}_Strategy_Summary.csv')
    pd.DataFrame(summary_rows).to_csv(summary_path, index=False)

    # Trade log
    trade_rows = []
    for name in strategies:
        for t in strategies[name][1]:
            trade_rows.append({
                'Strategy': name, 'Trade #': t['trade_num'],
                'Buy Date': t['buy_date'].strftime('%Y-%m-%d') if t['buy_date'] else '',
                'Buy Price': f"${t['buy_price']:.2f}",
                'Sell Date': t['sell_date'].strftime('%Y-%m-%d'),
                'Sell Price': f"${t['sell_price']:.2f}",
                'Shares': t['shares'],
                'Gross P&L': f"${t['gross_pnl']:,.2f}",
                'Return': f"{t['return_pct']:.2%}",
                'Result': 'WIN' if t['gross_pnl'] > 0 else 'LOSS',
            })
    trade_path = os.path.join(output_dir, f'{ticker_label}_Trade_Log.csv')
    pd.DataFrame(trade_rows).to_csv(trade_path, index=False)

    print(f"  CSV saved: {summary_path}")
    print(f"  CSV saved: {trade_path}")
    return summary_path, trade_path


# ============================================================
# CONSOLE REPORT
# ============================================================

def print_report(detection_results, strategies, rankings, ticker_label, capital):
    """Print formatted console report."""
    prices_n = len(list(strategies.values())[0][0])

    print("=" * 110)
    print(f"  {ticker_label} — TURNING POINT ANALYSIS REPORT")
    print(f"  Starting Capital: ${capital:,.0f} | Logic: BUY at troughs, SELL at peaks")
    print("=" * 110)

    # Turning points detected
    print("\n  TURNING POINTS DETECTED:")
    print("  " + "-" * 80)
    for name in [k for k in detection_results if not k.startswith('_')]:
        d = detection_results[name]
        if d.get('type') == 'scherman':
            sig_count = len(d.get('signals', []))
            print(f"  {name:<30} Divergence signals: {sig_count}")
        else:
            p = len(d['peaks'])
            t = len(d['troughs'])
            print(f"  {name:<30} Peaks: {p:>3}  |  Troughs: {t:>3}")

    # Strategy comparison
    print("\n  STRATEGY P&L COMPARISON:")
    print("  " + "-" * 108)
    header = f"  {'Strategy':<25} {'End Equity':>12} {'Total P&L':>12} {'Return':>8} {'Trades':>7} {'Win Rate':>9} {'Max DD':>8} {'Sharpe':>7}"
    print(header)
    print("  " + "-" * 108)

    for name in strategies:
        stats = strategies[name][2]
        print(f"  {name:<25} ${stats['final_eq']:>10,.0f} ${stats['total_pnl']:>10,.0f} "
              f"{stats['total_ret']:>7.1%} {stats['num_trades']:>6} {stats['win_rate']:>8.0%} "
              f"{stats['max_dd']:>7.1%} {stats['sharpe']:>6.2f}")

    # Method ranking
    if rankings:
        print(f"\n  METHOD RANKING:")
        print("  " + "-" * 80)
        for i, r in enumerate(rankings):
            print(f"  #{i+1}  {r['name']:<30} Score: {r['score']:.1f}  |  "
                  f"Precision: {r['precision']:.0f}%  |  Detections: {r['total']}")
        print(f"\n  BEST METHOD: {rankings[0]['name']} (Score: {rankings[0]['score']:.1f})")


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def run_full_analysis(
    ticker=None, csv_path=None, period='1y',
    start_date=None, end_date=None,
    capital=100000.0, output_dir=None
):
    """
    Run the complete Stock Turning Point Analysis pipeline.

    Parameters:
        ticker: Stock ticker symbol (e.g., 'AAPL'). Uses yfinance to download data.
        csv_path: Path to CSV file with Date/Close columns. Alternative to ticker.
        period: yfinance period ('1y', '2y', '6mo', etc.). Ignored if start/end given.
        start_date: Start date 'YYYY-MM-DD'. Overrides period.
        end_date: End date 'YYYY-MM-DD'. Overrides period.
        capital: Starting capital for backtests (default $100,000).
        output_dir: Directory for output files. Defaults to current directory.

    Returns:
        dict with keys: 'detection', 'strategies', 'rankings', 'files'
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tp_output')
    os.makedirs(output_dir, exist_ok=True)

    ticker_label = ticker.upper() if ticker else os.path.splitext(os.path.basename(csv_path or 'DATA'))[0]
    ticker_label = ticker_label.split('_')[0]  # Clean up label

    print(f"\n{'='*60}")
    print(f"  STOCK TURNING POINT ANALYSIS ENGINE")
    print(f"  Ticker: {ticker_label} | Capital: ${capital:,.0f}")
    print(f"{'='*60}\n")

    # Step 1: Load data
    print("[1/5] Loading data...")
    df = load_data(ticker=ticker, csv_path=csv_path, period=period,
                   start_date=start_date, end_date=end_date)
    prices = df['Close'].values
    dates = df['Date'].values
    print(f"  Loaded {len(df)} trading days: {df['Date'].iloc[0].strftime('%Y-%m-%d')} to {df['Date'].iloc[-1].strftime('%Y-%m-%d')}")

    # Step 2: Detect turning points
    print("\n[2/5] Running 7 detection methods + ensemble...")
    detection = detect_all_turning_points(prices, df=df)
    for name in [k for k in detection if not k.startswith('_')]:
        d = detection[name]
        if d.get('type') == 'scherman':
            sig_count = len(d.get('signals', []))
            print(f"  {name}: {sig_count} divergence signals")
        else:
            p = len(d['peaks'])
            t = len(d['troughs'])
            print(f"  {name}: {p} peaks, {t} troughs")

    # Step 3: Backtest all strategies
    print("\n[3/5] Backtesting all strategies...")
    strategies = backtest_all(detection, prices, dates, capital)
    for name in strategies:
        stats = strategies[name][2]
        print(f"  {name}: {stats['total_ret']:+.1%} return, {stats['num_trades']} trades")

    # Step 4: Evaluate and rank methods
    print("\n[4/5] Evaluating and ranking methods...")
    rankings = evaluate_methods(detection, prices, dates)

    # Step 5: Generate outputs
    print("\n[5/5] Generating outputs...")
    tp_chart, ret_chart = generate_charts(df, detection, strategies, output_dir, ticker_label)
    xlsx_path = generate_excel(strategies, output_dir, ticker_label, capital)
    summary_csv, trade_csv = generate_csv(strategies, rankings, output_dir, ticker_label)

    # Print full report
    print_report(detection, strategies, rankings, ticker_label, capital)

    files = {
        'turning_points_chart': tp_chart,
        'cumulative_returns_chart': ret_chart,
        'excel_workbook': xlsx_path,
        'summary_csv': summary_csv,
        'trade_log_csv': trade_csv,
    }

    print(f"\n{'='*60}")
    print(f"  ANALYSIS COMPLETE — Output files:")
    for k, v in files.items():
        print(f"    {k}: {v}")
    print(f"{'='*60}\n")

    return {
        'detection': detection,
        'strategies': strategies,
        'rankings': rankings,
        'files': files,
        'dataframe': df,
    }


# ============================================================
# CLI ENTRY POINT
# ============================================================

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Stock Turning Point Analysis Engine')
    parser.add_argument('--ticker', '-t', type=str, help='Stock ticker symbol (e.g., AAPL)')
    parser.add_argument('--csv', '-c', type=str, help='Path to CSV file with Date/Close columns')
    parser.add_argument('--period', '-p', type=str, default='1y', help='yfinance period (default: 1y)')
    parser.add_argument('--start', type=str, help='Start date YYYY-MM-DD')
    parser.add_argument('--end', type=str, help='End date YYYY-MM-DD')
    parser.add_argument('--capital', '-k', type=float, default=100000.0, help='Starting capital (default: 100000)')
    parser.add_argument('--output', '-o', type=str, default='.', help='Output directory')

    args = parser.parse_args()
    if not args.ticker and not args.csv:
        parser.error("Must provide either --ticker or --csv")

    run_full_analysis(
        ticker=args.ticker, csv_path=args.csv, period=args.period,
        start_date=args.start, end_date=args.end,
        capital=args.capital, output_dir=args.output,
    )
