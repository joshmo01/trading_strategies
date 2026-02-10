"""
P&L for all 6 Turning Point Detection Strategies + Ensemble + Buy & Hold
Trading Logic: BUY at detected troughs, SELL at detected peaks
Starting Capital: $100,000
"""
import pandas as pd
import numpy as np
from scipy.signal import argrelextrema
from scipy.ndimage import gaussian_filter1d
import pywt
import ruptures
import warnings
warnings.filterwarnings('ignore')

df = pd.read_csv('/sessions/fervent-dazzling-euler/mnt/uploads/AAPL_2023-01-01_to_2023-12-31_data.csv')
df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)
prices = df['Close'].values
dates = df['Date'].values
n = len(prices)
CAPITAL = 100000.0

# ════════════════════════════════════════════════════════════
# RECOMPUTE ALL 6 METHODS (same as turning_points.py)
# ════════════════════════════════════════════════════════════

# --- Method 1: Second Derivative ---
sigma = 8
smoothed = gaussian_filter1d(prices, sigma=sigma)
d1 = np.gradient(smoothed)
d2 = np.gradient(d1)
m1_peaks, m1_troughs = [], []
for i in range(1, len(d1)):
    if d1[i-1] > 0 and d1[i] <= 0 and d2[i] < 0: m1_peaks.append(i)
    elif d1[i-1] < 0 and d1[i] >= 0 and d2[i] > 0: m1_troughs.append(i)

# --- Method 2: RDP ---
def rdp(points, epsilon):
    if len(points) <= 2: return [0, len(points)-1]
    start, end = points[0], points[-1]
    line_vec = end - start
    line_len = np.linalg.norm(line_vec)
    if line_len == 0: return [0, len(points)-1]
    distances = np.abs(np.cross(line_vec, start - points)) / line_len
    max_idx = np.argmax(distances)
    if distances[max_idx] > epsilon:
        left = rdp(points[:max_idx+1], epsilon)
        right = [r + max_idx for r in rdp(points[max_idx:], epsilon)]
        return left[:-1] + right
    return [0, len(points)-1]

price_norm = (prices - prices.min()) / (prices.max() - prices.min())
points_2d = np.column_stack([np.arange(n)/n, price_norm])
rdp_indices = rdp(points_2d, 0.02)
m2_peaks, m2_troughs = [], []
for i, idx in enumerate(rdp_indices):
    if i == 0 or i == len(rdp_indices)-1: continue
    prev_idx, next_idx = rdp_indices[i-1], rdp_indices[i+1]
    if prices[idx] > prices[prev_idx] and prices[idx] > prices[next_idx]: m2_peaks.append(idx)
    elif prices[idx] < prices[prev_idx] and prices[idx] < prices[next_idx]: m2_troughs.append(idx)

# --- Method 3: Wavelet ---
pad_len = 2**int(np.ceil(np.log2(n))) - n
padded = np.pad(prices, (0, pad_len), mode='edge')
coeffs = pywt.wavedec(padded, 'db4', level=4)
for i in range(2, len(coeffs)): coeffs[i] = np.zeros_like(coeffs[i])
reconstructed = pywt.waverec(coeffs, 'db4')[:n]
m3_peaks_raw = argrelextrema(reconstructed, np.greater, order=15)[0]
m3_troughs_raw = argrelextrema(reconstructed, np.less, order=15)[0]

def map_to_actual(indices, prices, window=5, find_max=True):
    mapped = []
    for idx in indices:
        s, e = max(0, idx-window), min(len(prices), idx+window+1)
        mapped.append(s + (np.argmax(prices[s:e]) if find_max else np.argmin(prices[s:e])))
    return mapped

m3_peaks = map_to_actual(m3_peaks_raw, prices, 5, True)
m3_troughs = map_to_actual(m3_troughs_raw, prices, 5, False)

# --- Method 4: PELT ---
returns = np.diff(np.log(prices))
algo = ruptures.Pelt(model="rbf", min_size=10).fit(returns.reshape(-1,1))
change_points = [cp for cp in algo.predict(pen=1.5) if cp < n]
m4_peaks, m4_troughs = [], []
for cp in change_points:
    if cp <= 2 or cp >= n-2: continue
    lookback = min(10, cp)
    lookahead = min(10, n-cp)
    before = prices[cp] - prices[cp-lookback]
    after = prices[min(cp+lookahead, n-1)] - prices[cp]
    if before > 0 and after < 0: m4_peaks.append(cp)
    elif before < 0 and after > 0: m4_troughs.append(cp)

# --- Method 5: Bry-Boschan ---
def bry_boschan(prices, min_cycle=15, window=5):
    n = len(prices)
    sm = pd.Series(prices).rolling(13, center=True).mean().values
    for i in range(6): sm[i] = prices[i]; sm[-(i+1)] = prices[-(i+1)]
    candidates = []
    for i in range(window, n-window):
        w = sm[i-window:i+window+1]
        if sm[i] == np.nanmax(w): candidates.append((i, 'peak'))
        elif sm[i] == np.nanmin(w): candidates.append((i, 'trough'))
    if not candidates: return [], []
    alt = [candidates[0]]
    for c in candidates[1:]:
        if c[1] != alt[-1][1]: alt.append(c)
        elif c[1] == 'peak' and prices[c[0]] > prices[alt[-1][0]]: alt[-1] = c
        elif c[1] == 'trough' and prices[c[0]] < prices[alt[-1][0]]: alt[-1] = c
    filt = [alt[0]]
    for c in alt[1:]:
        if c[0] - filt[-1][0] >= min_cycle: filt.append(c)
        elif c[1] == 'peak' and prices[c[0]] > prices[filt[-1][0]]: filt[-1] = c
        elif c[1] == 'trough' and prices[c[0]] < prices[filt[-1][0]]: filt[-1] = c
    peaks, troughs = [], []
    for idx, tp in filt:
        s, e = max(0, idx-window), min(n, idx+window+1)
        if tp == 'peak': peaks.append(s + np.argmax(prices[s:e]))
        else: troughs.append(s + np.argmin(prices[s:e]))
    all_tp = sorted([(p, 'peak') for p in peaks] + [(t, 'trough') for t in troughs], key=lambda x: x[0])
    fp, ft = [], []
    for i, (idx, tp) in enumerate(all_tp):
        if i == 0: (fp if tp == 'peak' else ft).append(idx); continue
        amp = abs(prices[idx] - prices[all_tp[i-1][0]]) / prices[all_tp[i-1][0]] * 100
        if amp >= 3.0: (fp if tp == 'peak' else ft).append(idx)
    return fp, ft

m5_peaks, m5_troughs = bry_boschan(prices)

# --- Method 6: Menger Curvature ---
def menger_curvature(x, y, step=3):
    n = len(x); kappa = np.zeros(n)
    for i in range(step, n-step):
        x1,y1 = x[i-step],y[i-step]; x2,y2 = x[i],y[i]; x3,y3 = x[i+step],y[i+step]
        area = abs((x2-x1)*(y3-y1)-(x3-x1)*(y2-y1))/2
        d12 = np.sqrt((x2-x1)**2+(y2-y1)**2)
        d23 = np.sqrt((x3-x2)**2+(y3-y2)**2)
        d13 = np.sqrt((x3-x1)**2+(y3-y1)**2)
        denom = d12*d23*d13
        if denom > 0: kappa[i] = 4*area/denom
    return kappa

t_norm = np.arange(n)/n
p_norm = (smoothed-smoothed.min())/(smoothed.max()-smoothed.min())
curv = menger_curvature(t_norm, p_norm, step=5)
thresh = np.percentile(curv[curv>0], 85)
hc = np.where(curv > thresh)[0]
m6_peaks, m6_troughs = [], []
if len(hc) > 0:
    clusters, cur = [], [hc[0]]
    for i in range(1, len(hc)):
        if hc[i]-hc[i-1] <= 5: cur.append(hc[i])
        else: clusters.append(cur); cur = [hc[i]]
    clusters.append(cur)
    dd = np.gradient(smoothed)
    for cl in clusters:
        mi = cl[np.argmax(curv[cl])]
        if 0 < mi < n-1:
            if dd[mi-1] > 0 and dd[min(mi+1,n-1)] < 0: m6_peaks.append(mi)
            elif dd[mi-1] < 0 and dd[min(mi+1,n-1)] > 0: m6_troughs.append(mi)

# --- Ensemble Consensus ---
peak_score = np.zeros(n)
trough_score = np.zeros(n)
prox = 5
for methods, sc in [
    ([(m1_peaks,1.0),(m2_peaks,0.8),(m3_peaks,1.0),(m4_peaks,0.7),(m5_peaks,1.2),(m6_peaks,0.9)], peak_score),
    ([(m1_troughs,1.0),(m2_troughs,0.8),(m3_troughs,1.0),(m4_troughs,0.7),(m5_troughs,1.2),(m6_troughs,0.9)], trough_score)
]:
    for indices, w in methods:
        for idx in indices:
            for d in range(max(0,idx-prox), min(n,idx+prox+1)):
                sc[d] += w * np.exp(-0.5*(abs(d-idx)/2)**2)

ens_peaks = [p for p in argrelextrema(peak_score, np.greater, order=10)[0] if peak_score[p] >= 2.0]
ens_troughs = [t for t in argrelextrema(trough_score, np.greater, order=10)[0] if trough_score[t] >= 2.0]

# ════════════════════════════════════════════════════════════
# TRADE ENGINE: Buy at troughs, Sell at peaks
# ════════════════════════════════════════════════════════════

def generate_trades(peaks, troughs, prices, dates):
    """Interleave troughs (buys) and peaks (sells) chronologically."""
    events = sorted(
        [(t, 'BUY', prices[t], dates[t]) for t in troughs] +
        [(p, 'SELL', prices[p], dates[p]) for p in peaks],
        key=lambda x: x[0]
    )
    # Enforce alternation: must start with BUY, then alternate
    trades = []
    looking_for = 'BUY'
    for idx, action, price, date in events:
        if action == looking_for:
            trades.append({'idx': idx, 'action': action, 'price': price, 'date': pd.Timestamp(date)})
            looking_for = 'SELL' if action == 'BUY' else 'BUY'
    return trades

def compute_pnl(trade_sequence, prices, dates, capital=CAPITAL):
    """Compute full P&L from alternating BUY/SELL sequence."""
    cash = capital
    shares = 0
    cost_basis = 0
    position = 'CASH'
    cum_realized = 0.0
    trade_num = 0
    completed_trades = []

    # Build position array (1=long, 0=cash) for each day
    positions = np.zeros(len(prices))
    buy_idx = None
    for t in trade_sequence:
        if t['action'] == 'BUY':
            buy_idx = t['idx']
        elif t['action'] == 'SELL' and buy_idx is not None:
            positions[buy_idx:t['idx']+1] = 1
            buy_idx = None
    if buy_idx is not None:
        positions[buy_idx:] = 1  # Still holding at end

    # Daily simulation
    daily_rows = []
    for i in range(len(prices)):
        prev_pos = positions[i-1] if i > 0 else 0
        curr_pos = positions[i]
        action = ''
        trade_pnl = 0.0

        if curr_pos == 1 and prev_pos == 0:
            action = 'BUY'
            trade_num += 1
            shares = int(cash / prices[i])
            cost_basis = shares * prices[i]
            cash -= cost_basis
            position = 'LONG'
        elif curr_pos == 0 and prev_pos == 1:
            action = 'SELL'
            proceeds = shares * prices[i]
            trade_pnl = proceeds - cost_basis
            cum_realized += trade_pnl
            completed_trades.append({
                'trade_num': trade_num,
                'buy_date': None,  # filled below
                'sell_date': pd.Timestamp(dates[i]),
                'buy_price': cost_basis / shares if shares > 0 else 0,
                'sell_price': prices[i],
                'shares': shares,
                'gross_pnl': trade_pnl,
                'return_pct': (proceeds / cost_basis - 1) if cost_basis > 0 else 0,
            })
            cash += proceeds
            shares = 0
            cost_basis = 0
            position = 'CASH'

        mkt_value = shares * prices[i]
        unrealized = mkt_value - cost_basis if shares > 0 else 0
        total_equity = cash + mkt_value
        prev_equity = daily_rows[-1]['Total_Equity'] if daily_rows else capital
        daily_ret = ((total_equity / prev_equity) - 1) * 100 if prev_equity > 0 else 0

        daily_rows.append({
            'Date': pd.Timestamp(dates[i]),
            'Close': prices[i],
            'Action': action,
            'Trade_#': trade_num if action else '',
            'Position': position,
            'Shares': shares,
            'Cost_Basis': cost_basis if shares > 0 else 0,
            'Mkt_Value': mkt_value,
            'Unrealized_PnL': unrealized,
            'Trade_PnL': trade_pnl if action == 'SELL' else '',
            'Cum_Realized_PnL': cum_realized,
            'Cash': cash,
            'Total_Equity': total_equity,
            'Daily_Return_Pct': daily_ret,
        })

    # Fill buy_dates in completed trades
    buy_dates = [t for t in trade_sequence if t['action'] == 'BUY']
    for i, ct in enumerate(completed_trades):
        if i < len(buy_dates):
            ct['buy_date'] = buy_dates[i]['date']

    return pd.DataFrame(daily_rows), completed_trades

# ════════════════════════════════════════════════════════════
# RUN ALL STRATEGIES
# ════════════════════════════════════════════════════════════

strategies = {}
method_configs = [
    ('1. 2nd Derivative', m1_peaks, m1_troughs),
    ('2. RDP Linearization', m2_peaks, m2_troughs),
    ('3. Wavelet (db4)', m3_peaks, m3_troughs),
    ('4. PELT Change Point', m4_peaks, m4_troughs),
    ('5. Bry-Boschan', m5_peaks, m5_troughs),
    ('6. Menger Curvature', m6_peaks, m6_troughs),
    ('7. Ensemble Consensus', ens_peaks, ens_troughs),
]

# Buy & Hold
bh_positions = np.ones(n)
bh_daily = []
bh_shares = int(CAPITAL / prices[0])
bh_cost = bh_shares * prices[0]
bh_cash = CAPITAL - bh_cost
for i in range(n):
    mv = bh_shares * prices[i]
    eq = bh_cash + mv
    prev_eq = bh_daily[-1]['Total_Equity'] if bh_daily else CAPITAL
    bh_daily.append({
        'Date': pd.Timestamp(dates[i]), 'Close': prices[i],
        'Action': 'BUY' if i == 0 else ('SELL' if i == n-1 else ''),
        'Trade_#': 1 if i == 0 or i == n-1 else '',
        'Position': 'LONG', 'Shares': bh_shares,
        'Cost_Basis': bh_cost, 'Mkt_Value': mv,
        'Unrealized_PnL': mv - bh_cost,
        'Trade_PnL': mv - bh_cost if i == n-1 else '',
        'Cum_Realized_PnL': 0 if i < n-1 else mv - bh_cost,
        'Cash': bh_cash, 'Total_Equity': eq,
        'Daily_Return_Pct': ((eq/prev_eq)-1)*100 if prev_eq > 0 else 0,
    })
bh_df = pd.DataFrame(bh_daily)
bh_completed = [{
    'trade_num': 1, 'buy_date': pd.Timestamp(dates[0]), 'sell_date': pd.Timestamp(dates[-1]),
    'buy_price': prices[0], 'sell_price': prices[-1], 'shares': bh_shares,
    'gross_pnl': bh_shares * (prices[-1] - prices[0]),
    'return_pct': prices[-1]/prices[0] - 1,
}]
strategies['Buy & Hold'] = (bh_df, bh_completed)

for name, peaks, troughs in method_configs:
    seq = generate_trades(peaks, troughs, prices, dates)
    daily_df, completed = compute_pnl(seq, prices, dates)
    strategies[name] = (daily_df, completed)

# ════════════════════════════════════════════════════════════
# PRINT SUMMARY
# ════════════════════════════════════════════════════════════

print("=" * 110)
print("  AAPL 2023 — TURNING POINT STRATEGY P&L COMPARISON")
print("  Starting Capital: $100,000 | Logic: BUY at troughs, SELL at peaks")
print("=" * 110)
print()

header = f"{'Strategy':<25} {'End Equity':>12} {'Total P&L':>12} {'Return':>8} {'Trades':>7} {'Winners':>8} {'Win Rate':>9} {'Max DD':>8} {'Sharpe':>7}"
print(header)
print("─" * 110)

summary_data = []
for name in ['Buy & Hold'] + [mc[0] for mc in method_configs]:
    daily_df, completed = strategies[name]
    final_eq = daily_df['Total_Equity'].iloc[-1]
    total_pnl = final_eq - CAPITAL
    total_ret = (final_eq / CAPITAL - 1)
    num_trades = len(completed)
    winners = sum(1 for t in completed if t['gross_pnl'] > 0)
    losers = num_trades - winners
    win_rate = winners / num_trades if num_trades > 0 else 0
    daily_rets = daily_df['Daily_Return_Pct'] / 100
    vol = daily_rets.std() * np.sqrt(252)
    sharpe = (daily_rets.mean() / daily_rets.std()) * np.sqrt(252) if daily_rets.std() > 0 else 0
    peak = daily_df['Total_Equity'].cummax()
    max_dd = ((daily_df['Total_Equity'] - peak) / peak).min()
    days_long = daily_df[daily_df['Position'] == 'LONG'].shape[0]
    exposure = days_long / n

    best_t = max([t['return_pct'] for t in completed]) if completed else 0
    worst_t = min([t['return_pct'] for t in completed]) if completed else 0
    avg_t = np.mean([t['return_pct'] for t in completed]) if completed else 0

    summary_data.append({
        'name': name, 'final_eq': final_eq, 'total_pnl': total_pnl,
        'total_ret': total_ret, 'num_trades': num_trades,
        'winners': winners, 'losers': losers, 'win_rate': win_rate,
        'best_trade': best_t, 'worst_trade': worst_t, 'avg_trade': avg_t,
        'vol': vol, 'sharpe': sharpe, 'max_dd': max_dd,
        'exposure': exposure, 'days_long': days_long,
    })

    flag = " ◀ BEST" if name != 'Buy & Hold' and total_pnl == max(
        s['total_pnl'] for s in summary_data if s['name'] != 'Buy & Hold'
    ) else ""
    print(f"  {name:<23} ${final_eq:>10,.0f} ${total_pnl:>10,.0f} {total_ret:>7.1%} {num_trades:>6} {winners:>7} {win_rate:>8.0%} {max_dd:>7.1%} {sharpe:>6.2f}{flag}")

print()
# Trade details
for name in [mc[0] for mc in method_configs] + ['7. Ensemble Consensus']:
    if name == '7. Ensemble Consensus':
        name_key = '7. Ensemble Consensus'
    else:
        name_key = name
    if name_key not in strategies:
        continue
    daily_df, completed = strategies[name_key]
    print(f"\n  ── {name_key} Trade Log ──")
    if not completed:
        print("     No trades generated.")
        continue
    for t in completed:
        result = "WIN " if t['gross_pnl'] > 0 else "LOSS"
        bd = t['buy_date'].strftime('%b %d') if t['buy_date'] else '?'
        sd = t['sell_date'].strftime('%b %d') if t['sell_date'] else '?'
        print(f"     #{t['trade_num']:>2}  {bd} → {sd}  |  Buy ${t['buy_price']:.2f} → Sell ${t['sell_price']:.2f}  |  "
              f"Shares: {t['shares']}  |  P&L: ${t['gross_pnl']:>+10,.2f}  ({t['return_pct']:>+6.2%})  {result}")

# ════════════════════════════════════════════════════════════
# BUILD EXCEL WORKBOOK
# ════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill('solid', fgColor='1a237e')
title_font = Font(name='Arial', bold=True, size=14, color='1a237e')
num_font = Font(name='Arial', size=10)
green_font = Font(name='Arial', size=10, color='008000', bold=True)
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
buy_fill = PatternFill('solid', fgColor='E8F5E9')
sell_fill = PatternFill('solid', fgColor='FFEBEE')
best_fill = PatternFill('solid', fgColor='C8E6C9')
total_fill = PatternFill('solid', fgColor='E3F2FD')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
)
money_fmt = '$#,##0.00'
pnl_fmt = '$#,##0.00;[Red]($#,##0.00);"-"'

# ═══ SUMMARY SHEET ═══
ws = wb.active
ws.title = 'P&L Summary'
ws.sheet_properties.tabColor = '1a237e'

ws['A1'] = 'AAPL 2023 — Turning Point Strategy P&L Comparison'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='1a237e')
ws['A2'] = 'Starting Capital: $100,000 | Logic: BUY at detected troughs, SELL at detected peaks'
ws['A2'].font = Font(name='Arial', size=10, color='666666')

all_names = ['Buy & Hold'] + [mc[0] for mc in method_configs]
headers = ['Metric'] + all_names
r = 4
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

metrics_order = [
    ('Starting Capital', lambda s: CAPITAL, money_fmt),
    ('Ending Equity', lambda s: s['final_eq'], money_fmt),
    ('Total P&L ($)', lambda s: s['total_pnl'], pnl_fmt),
    ('Total Return (%)', lambda s: s['total_ret'], '0.00%'),
    ('Number of Trades', lambda s: s['num_trades'], '0'),
    ('Winning Trades', lambda s: s['winners'], '0'),
    ('Losing Trades', lambda s: s['losers'], '0'),
    ('Win Rate (%)', lambda s: s['win_rate'], '0.0%'),
    ('Best Trade (%)', lambda s: s['best_trade'], '0.00%'),
    ('Worst Trade (%)', lambda s: s['worst_trade'], '0.00%'),
    ('Avg Trade Return (%)', lambda s: s['avg_trade'], '0.00%'),
    ('Annualized Volatility (%)', lambda s: s['vol'], '0.00%'),
    ('Sharpe Ratio', lambda s: s['sharpe'], '0.00'),
    ('Max Drawdown (%)', lambda s: s['max_dd'], '0.00%'),
    ('Market Exposure (%)', lambda s: s['exposure'], '0.0%'),
    ('Days in Market', lambda s: s['days_long'], '0'),
    ('Days in Cash', lambda s: n - s['days_long'], '0'),
]

for mi, (metric_name, func, fmt) in enumerate(metrics_order):
    row = r + 1 + mi
    cell = ws.cell(row=row, column=1, value=metric_name)
    cell.font = Font(name='Arial', size=10, bold=True); cell.border = thin_border

    vals = []
    for ci, name in enumerate(all_names):
        sd = [s for s in summary_data if s['name'] == name][0]
        val = func(sd)
        vals.append(val)
        cell = ws.cell(row=row, column=ci+2, value=val)
        cell.number_format = fmt; cell.font = num_font
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

        if metric_name in ['Total P&L ($)', 'Total Return (%)']:
            cell.font = green_font if val > 0 else (red_font if val < 0 else num_font)
        if metric_name == 'Ending Equity':
            cell.fill = total_fill

    # Highlight best (skip Buy & Hold for comparison)
    if metric_name not in ['Starting Capital', 'Days in Cash', 'Days in Market']:
        strat_vals = vals[1:]  # exclude buy & hold
        if metric_name in ['Losing Trades', 'Worst Trade (%)', 'Max Drawdown (%)', 'Annualized Volatility (%)']:
            best_i = strat_vals.index(max(strat_vals))
        else:
            best_i = strat_vals.index(max(strat_vals))
        ws.cell(row=row, column=best_i+3).fill = best_fill  # +3 because col 1=metric, col 2=B&H

ws.column_dimensions['A'].width = 28
for c in range(2, len(all_names)+2):
    ws.column_dimensions[get_column_letter(c)].width = 20

# ═══ TRADE LOG SHEET ═══
tl = wb.create_sheet('Trade Log')
tl.sheet_properties.tabColor = 'FF7043'
tl['A1'] = 'Trade-by-Trade P&L — All Strategies'
tl['A1'].font = title_font

trade_headers = ['Strategy', 'Trade #', 'Buy Date', 'Buy Price', 'Sell Date', 'Sell Price',
                 'Shares', 'Gross P&L ($)', 'Return (%)', 'Holding Days', 'Result']
tr = 3
for c, h in enumerate(trade_headers, 1):
    cell = tl.cell(row=tr, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

for name in all_names:
    daily_df, completed = strategies[name]
    for t in completed:
        tr += 1
        days = (t['sell_date'] - t['buy_date']).days if t['buy_date'] else 0
        result = 'WIN' if t['gross_pnl'] > 0 else 'LOSS'
        vals = [name, t['trade_num'],
                t['buy_date'].strftime('%Y-%m-%d') if t['buy_date'] else '',
                t['buy_price'], t['sell_date'].strftime('%Y-%m-%d'),
                t['sell_price'], t['shares'], t['gross_pnl'], t['return_pct'], days, result]
        fmts = [None, '0', None, money_fmt, None, money_fmt, '#,##0', pnl_fmt, '0.00%', '0', None]

        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = tl.cell(row=tr, column=c, value=val)
            if fmt: cell.number_format = fmt
            cell.font = num_font; cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if c == 11: cell.font = green_font if val == 'WIN' else red_font

        fill = buy_fill if t['gross_pnl'] > 0 else sell_fill
        for c in range(1, 12): tl.cell(row=tr, column=c).fill = fill

for c in range(1, 12):
    tl.column_dimensions[get_column_letter(c)].width = max(14, len(trade_headers[c-1])+4)
tl.column_dimensions['A'].width = 24
tl.freeze_panes = 'A4'

# ═══ DAILY P&L SHEETS ═══
tab_colors = ['4CAF50', '2196F3', 'AB47BC', 'FF9800', '009688', 'E91E63', 'FF5722', '3F51B5']
daily_headers = ['Date', 'Close', 'Action', 'Trade #', 'Position', 'Shares',
                 'Cost Basis', 'Mkt Value', 'Unrealized P&L', 'Trade P&L',
                 'Cum Realized P&L', 'Cash', 'Total Equity', 'Daily Return %']
col_fmts = [None, money_fmt, None, None, None, '#,##0',
            money_fmt, money_fmt, pnl_fmt, pnl_fmt, pnl_fmt, money_fmt, money_fmt, '0.00%']

for si, name in enumerate(all_names):
    daily_df, _ = strategies[name]
    short_name = name[:25]
    ws = wb.create_sheet(short_name)
    ws.sheet_properties.tabColor = tab_colors[si % len(tab_colors)]
    ws['A1'] = f'{name} — Daily P&L'
    ws['A1'].font = title_font

    for c, h in enumerate(daily_headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin_border

    for i, row in daily_df.iterrows():
        rr = i + 4
        values = [
            row['Date'].strftime('%Y-%m-%d'), row['Close'], row['Action'],
            row['Trade_#'], row['Position'], row['Shares'],
            row['Cost_Basis'], row['Mkt_Value'], row['Unrealized_PnL'],
            row['Trade_PnL'] if row['Trade_PnL'] != '' else '',
            row['Cum_Realized_PnL'], row['Cash'], row['Total_Equity'],
            row['Daily_Return_Pct'] / 100
        ]
        for c, (val, fmt) in enumerate(zip(values, col_fmts), 1):
            cell = ws.cell(row=rr, column=c, value=val)
            if fmt and val != '': cell.number_format = fmt
            cell.font = num_font; cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        if row['Action'] == 'BUY':
            for c in range(1, 15): ws.cell(row=rr, column=c).fill = buy_fill
            ws.cell(row=rr, column=3).font = green_font
        elif row['Action'] == 'SELL':
            for c in range(1, 15): ws.cell(row=rr, column=c).fill = sell_fill
            ws.cell(row=rr, column=3).font = red_font

    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = max(13, len(daily_headers[c-1])+3)
    ws.freeze_panes = 'A4'

out = '/sessions/fervent-dazzling-euler/mnt/outputs/AAPL_2023_TurningPoint_PnL.xlsx'
wb.save(out)
print(f"\n✓ Excel workbook saved: {out}")
